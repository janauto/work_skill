#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill

from workbook_utils import bool_from_value


HEADER_FILL = PatternFill("solid", fgColor="FF4F81BD")
TYPE_FILL = PatternFill("solid", fgColor="FFD9EAF7")
CATEGORY_FILL = PatternFill("solid", fgColor="FFFFFF99")
HEADER_FONT = Font(bold=True, color="FFFFFFFF")
RED_FONT = Font(name="Microsoft YaHei", size=11, color="FFFF0000")
BASE_FONT = Font(name="Microsoft YaHei", size=11)
BASE_BOLD_FONT = Font(name="Microsoft YaHei", size=11, bold=True)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")

RISK_CATEGORIES = {
    "质量/故障",
    "质量故障",
    "品质",
    "故障",
    "兼容性",
    "兼容性/连接问题",
    "兼容连接",
    "噪音",
    "噪音问题",
    "功能需求",
}
MACRO_THEME_MAP = {
    "音质": "声音相关",
    "音质/性能体验": "声音相关",
    "音质体验": "声音相关",
    "噪音": "声音相关",
    "噪音问题": "声音相关",
    "质量/故障": "品质相关",
    "质量故障": "品质相关",
    "品质": "品质相关",
    "故障": "品质相关",
    "功能": "功能相关",
    "功能需求": "功能相关",
    "接口/连接": "功能相关",
    "兼容性": "兼容与适配",
    "兼容性/连接问题": "兼容与适配",
    "兼容连接": "兼容与适配",
    "用户": "用户与认知",
    "用户原因": "用户与认知",
    "页面/预期不符": "体验与认知",
    "页面预期不符": "体验与认知",
    "体验": "体验与认知",
    "价格": "价格与竞品",
    "价格/替代品/竞品": "价格与竞品",
    "价格竞品": "价格与竞品",
    "降价": "价格与竞品",
    "物流": "交付与包装",
    "物流包装": "交付与包装",
    "包装/配件/版本问题": "交付与包装",
}
PASTEL_FILL_COLORS = [
    "FFF7E7",
    "FFFDEB",
    "FFEFE4",
    "FFEFD6",
    "FFE8F3",
    "FFEAF4",
    "FFE8F0",
    "FFEDEDED",
]


def find_input_sheet(wb, requested_sheet=None):
    if requested_sheet:
        return wb[requested_sheet]
    for name in ("TaggedComments", "CleanedComments"):
        if name in wb.sheetnames:
            return wb[name]
    return wb[wb.sheetnames[0]]


def read_metadata(wb):
    if "Metadata" not in wb.sheetnames:
        return {}
    ws = wb["Metadata"]
    metadata = {}
    for row_idx in range(2, ws.max_row + 1):
        key = ws.cell(row_idx, 1).value
        value = ws.cell(row_idx, 2).value
        if key:
            metadata[str(key)] = value
    return metadata


def row_dicts(ws):
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    for row_idx in range(2, ws.max_row + 1):
        row = {headers[col - 1]: ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)}
        if not any(value not in (None, "") for value in row.values()):
            continue
        yield row


def normalize_theme(level_1):
    if not level_1:
        return "未归类"
    return MACRO_THEME_MAP.get(level_1, level_1)


def parse_date(value):
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def get_time_value(row):
    for key in ("return_time", "退货时间", "return_date", "date"):
        if row.get(key):
            return row.get(key)
    return None


def build_trend_counter(valid_rows):
    dates = [parse_date(get_time_value(row)) for row in valid_rows]
    dates = [dt for dt in dates if dt is not None]
    if not dates:
        return Counter(), None

    unique_months = {dt.strftime("%Y-%m") for dt in dates}
    granularity = "month" if len(unique_months) >= 3 else "day"
    counter = Counter(
        dt.strftime("%Y-%m") if granularity == "month" else dt.strftime("%Y-%m-%d") for dt in dates
    )
    return counter, granularity


def summarize(rows):
    valid_rows = [row for row in rows if bool_from_value(row.get("is_valid_feedback", True))]
    tagged_rows = [row for row in valid_rows if row.get("level_1")]
    chain_counter = Counter()
    type_counter = Counter()
    subissue_counter = defaultdict(Counter)
    severity_counter = Counter()
    macro_counter = Counter()

    for row in tagged_rows:
        l1 = row.get("level_1") or "未分类"
        l2 = row.get("level_2") or ""
        l3 = row.get("level_3") or ""
        chain_counter[(l1, l2, l3)] += 1
        type_counter[l1] += 1
        macro_counter[normalize_theme(l1)] += 1
        if l2:
            subissue_counter[l1][l2] += 1
        severity = row.get("severity") or "unknown"
        severity_counter[severity] += 1

    trend_counter, trend_granularity = build_trend_counter(valid_rows)
    return (
        valid_rows,
        tagged_rows,
        chain_counter,
        type_counter,
        subissue_counter,
        severity_counter,
        macro_counter,
        trend_counter,
        trend_granularity,
    )


def build_summary_text(product, metadata, rows, valid_rows, type_counter, subissue_counter, macro_counter, trend_counter, trend_granularity):
    total_rows = len(rows)
    total_valid = len(valid_rows)
    matched_rows = int(metadata.get("matched_rows_before_dedupe", total_rows) or total_rows)
    duplicate_rows = int(metadata.get("duplicate_rows_dropped", 0) or 0)
    invalid_rows = int(metadata.get("invalid_rows_dropped", max(matched_rows - total_rows - duplicate_rows, 0)) or 0)

    if not total_valid:
        text = (
            f"总结：\n本次聚焦 {product}，共匹配到 {matched_rows} 条记录，但当前口径下暂无可用于汇总的有效用户反馈。"
        )
        return text, [product, "有效用户反馈"], (4, len(text))

    top_three = type_counter.most_common(3)
    top_text = "、".join(f"{name}{count / total_valid:.2%}" for name, count in top_three)
    focus_categories = " 和 ".join(name for name, _ in top_three[:2]) if len(top_three) >= 2 else top_three[0][0]

    overview = (
        f"本次聚焦 {product}，共匹配到 {matched_rows} 条记录，去重清洗后保留 {total_rows} 条，"
        f"其中有效反馈 {total_valid} 条。综合来看，当前用户对 {product} 的核心问题主要集中在 {focus_categories}。"
        f"Top3 为 {top_text}。"
    )

    lines = ["总结：", overview]
    highlight_terms = [product, "有效反馈"]
    highlight_terms.extend(name for name, _ in top_three)
    highlight_terms.extend(f"{count / total_valid:.2%}" for _, count in top_three)

    if invalid_rows:
        lines.append(
            f"说明：无评论或无效评论 {invalid_rows} 条未纳入有效反馈统计，重复记录额外剔除 {duplicate_rows} 条。"
        )
        highlight_terms.extend(["无评论或无效评论", "未纳入有效反馈统计"])
    elif duplicate_rows:
        lines.append(f"说明：重复记录剔除 {duplicate_rows} 条。")

    if trend_counter:
        peak_period, peak_count = trend_counter.most_common(1)[0]
        trend_name = "月份" if trend_granularity == "month" else "日期"
        lines.append(f"时间趋势上，{peak_period} 为有效反馈峰值{trend_name}，占全部有效反馈 {peak_count / total_valid:.2%}。")
        highlight_terms.extend([peak_period, "峰值", f"{peak_count / total_valid:.2%}"])

    lines.append("")

    ranked = sorted(
        type_counter.items(),
        key=lambda item: (item[0] not in RISK_CATEGORIES, -item[1], item[0]),
    )
    for index, (category, count) in enumerate(ranked[:4], start=1):
        pct = f"{count / total_valid:.2%}"
        top_subissues = "、".join(name for name, _ in subissue_counter[category].most_common(3) if name) or "需结合原评论进一步复核"
        lines.append(f"{index}、{category}（{pct}）")
        lines.append(f"主要问题集中在 {top_subissues}。")
        lines.append("")
        highlight_terms.extend([category, pct])
        highlight_terms.extend(name for name, _ in subissue_counter[category].most_common(3) if name)

    converged = "、".join(f"{name}{count / total_valid:.2%}" for name, count in macro_counter.most_common(4))
    lines.append(f"收敛归纳：可以进一步归并为 {converged}")
    highlight_terms.extend(name for name, _ in macro_counter.most_common(4))
    if any(name in RISK_CATEGORIES for name in type_counter):
        lines.append("风险提示：优先排查故障 兼容性 功率 噪音等高风险问题。")
        highlight_terms.extend(["风险提示", "故障", "兼容性", "功率", "噪音"])

    text = "\n".join(lines).strip()
    bold_start = text.find(overview)
    bold_span = (bold_start, bold_start + len(overview))
    highlight_terms = [term for term in dict.fromkeys(highlight_terms) if term]
    return text, highlight_terms, bold_span


def build_rich_text(text, red_terms, bold_span):
    matches = []
    for term in sorted(red_terms, key=len, reverse=True):
        for match in re.finditer(re.escape(term), text):
            start, end = match.span()
            if any(start < exist_end and end > exist_start for exist_start, exist_end in matches):
                continue
            matches.append((start, end))
    matches.sort()

    boundaries = {0, len(text), bold_span[0], bold_span[1]}
    for start, end in matches:
        boundaries.add(start)
        boundaries.add(end)
    ordered = sorted(boundaries)

    rich = CellRichText()
    for start, end in zip(ordered, ordered[1:]):
        if start == end:
            continue
        segment = text[start:end]
        if not segment:
            continue
        bold = bold_span[0] <= start and end <= bold_span[1]
        red = any(match_start <= start and end <= match_end for match_start, match_end in matches)
        if not bold and not red:
            rich.append(segment)
            continue
        font = InlineFont(
            rFont="Microsoft YaHei",
            sz=11,
            b=True if bold else None,
            color="FFFF0000" if red else None,
        )
        rich.append(TextBlock(font, segment))
    return rich


def category_fill_map(categories):
    mapping = {}
    for index, category in enumerate(categories):
        mapping[category] = PatternFill("solid", fgColor=PASTEL_FILL_COLORS[index % len(PASTEL_FILL_COLORS)])
    return mapping


def apply_font_and_alignment(ws, min_row=1, max_col=None):
    max_col = max_col or ws.max_column
    for row in ws.iter_rows(min_row=min_row, max_col=max_col):
        for cell in row:
            if cell.row == 1 or cell.font == HEADER_FONT:
                continue
            if cell.value is not None:
                cell.font = Font(
                    name="Microsoft YaHei",
                    size=11,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    color=cell.font.color.rgb if getattr(cell.font.color, "type", None) == "rgb" else None,
                )
                if cell.alignment != WRAP_CENTER:
                    cell.alignment = WRAP_TOP


def style_headers(ws, cells, fill):
    for coord in cells:
        cell = ws[coord]
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = WRAP_CENTER


def style_percentage_column(ws, column_letter, start_row, end_row):
    for row_idx in range(start_row, end_row + 1):
        cell = ws[f"{column_letter}{row_idx}"]
        if isinstance(cell.value, (int, float)):
            cell.number_format = "0.00%"
            cell.font = RED_FONT


def copy_source_sheet(source_ws, output_wb, title):
    copied = output_wb.create_sheet(title)
    for row in source_ws.iter_rows():
        copied.append([cell.value for cell in row])
    copied.freeze_panes = "A2"
    if copied.max_row >= 1:
        for cell in copied[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = WRAP_CENTER
    for row in copied.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP_TOP
    for column_cells in copied.columns:
        letter = column_cells[0].column_letter
        header = str(column_cells[0].value or "")
        if any(key in header for key in ("评论", "翻译", "comment", "remark")):
            copied.column_dimensions[letter].width = 42
        elif "时间" in header:
            copied.column_dimensions[letter].width = 14
        elif any(key in header for key in ("分类", "归因", "product_name", "source_sheet")):
            copied.column_dimensions[letter].width = 20
        else:
            copied.column_dimensions[letter].width = 14
    apply_font_and_alignment(copied, min_row=2)
    return copied


def write_source_classification_block(summary, chain_counter, total_valid, fills):
    summary.merge_cells("A1:E1")
    summary["A1"] = "源分类：按大类到小类逐层筛选"
    style_headers(summary, ["A1"], HEADER_FILL)
    headers = ["一级分类", "二级分类", "三级分类", "数量", "占比"]
    for offset, header in enumerate(headers, start=1):
        summary.cell(2, offset).value = header
    style_headers(summary, ["A2", "B2", "C2", "D2", "E2"], TYPE_FILL)

    grouped = defaultdict(list)
    for chain, count in chain_counter.items():
        grouped[chain[0]].append((chain, count))

    row_idx = 3
    for category, items in sorted(grouped.items(), key=lambda item: (-sum(count for _, count in item[1]), item[0])):
        first_row = row_idx
        fill = fills.get(category, CATEGORY_FILL)
        for item_index, (chain, count) in enumerate(sorted(items, key=lambda item: (-item[1], item[0]))):
            summary.cell(row_idx, 1).value = category if item_index == 0 else None
            summary.cell(row_idx, 2).value = chain[1]
            summary.cell(row_idx, 3).value = chain[2]
            summary.cell(row_idx, 4).value = count
            summary.cell(row_idx, 5).value = count / total_valid
            for col_idx in range(1, 6):
                summary.cell(row_idx, col_idx).fill = fill
            row_idx += 1
        if row_idx - 1 > first_row:
            summary.merge_cells(start_row=first_row, start_column=1, end_row=row_idx - 1, end_column=1)

    style_percentage_column(summary, "E", 3, row_idx - 1)
    return row_idx


def write_sorted_block(summary, subissue_counter, total_valid, fills):
    summary.merge_cells("G1:J1")
    summary["G1"] = "按降序排列：先看高频小类"
    style_headers(summary, ["G1"], HEADER_FILL)
    headers = ["上级分类", "小类", "数量", "占比"]
    for offset, header in enumerate(headers, start=7):
        summary.cell(2, offset).value = header
    style_headers(summary, ["G2", "H2", "I2", "J2"], TYPE_FILL)

    flat = []
    for parent, child_counter in subissue_counter.items():
        for child, count in child_counter.items():
            flat.append((parent, child, count, count / total_valid))
    flat.sort(key=lambda item: (-item[2], item[0], item[1]))

    row_idx = 3
    for parent, child, count, pct in flat:
        summary.cell(row_idx, 7).value = parent
        summary.cell(row_idx, 8).value = child
        summary.cell(row_idx, 9).value = count
        summary.cell(row_idx, 10).value = pct
        fill = fills.get(parent, CATEGORY_FILL)
        for col_idx in range(7, 11):
            summary.cell(row_idx, col_idx).fill = fill
        row_idx += 1
    style_percentage_column(summary, "J", 3, row_idx - 1)
    return row_idx


def write_converged_block(summary, type_counter, macro_counter, total_valid, fills):
    summary.merge_cells("L1:P1")
    summary["L1"] = "收敛归纳：按类别再次分大类"
    style_headers(summary, ["L1"], HEADER_FILL)
    headers = ["原分类", "收敛大类", "数量", "占比", "说明"]
    for offset, header in enumerate(headers, start=12):
        summary.cell(2, offset).value = header
    style_headers(summary, ["L2", "M2", "N2", "O2", "P2"], TYPE_FILL)

    row_idx = 3
    for category, count in type_counter.most_common():
        macro = normalize_theme(category)
        summary.cell(row_idx, 12).value = category
        summary.cell(row_idx, 13).value = macro
        summary.cell(row_idx, 14).value = count
        summary.cell(row_idx, 15).value = count / total_valid
        summary.cell(row_idx, 16).value = f"归并到 {macro}"
        fill = fills.get(category, CATEGORY_FILL)
        for col_idx in range(12, 17):
            summary.cell(row_idx, col_idx).fill = fill
        row_idx += 1

    summary.merge_cells(start_row=3, start_column=18, end_row=6, end_column=20)
    macro_text = "\n".join(f"{name}\n{count}条\n{count / total_valid:.2%}" for name, count in macro_counter.most_common(4))
    summary.cell(3, 18).value = macro_text
    summary.cell(3, 18).alignment = WRAP_CENTER
    summary.cell(3, 18).font = BASE_BOLD_FONT

    style_percentage_column(summary, "O", 3, row_idx - 1)
    return row_idx


def add_time_trend(summary, trend_counter, trend_granularity, total_valid, start_row):
    if not trend_counter or len(trend_counter) < 2:
        return

    period_label = "月份" if trend_granularity == "month" else "日期"
    summary.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=3)
    summary.cell(start_row, 1).value = f"退货趋势：按{period_label}有效反馈"
    style_headers(summary, [f"A{start_row}"], HEADER_FILL)

    summary.cell(start_row + 1, 1).value = period_label
    summary.cell(start_row + 1, 2).value = "有效反馈数"
    summary.cell(start_row + 1, 3).value = "占比"
    style_headers(summary, [f"A{start_row + 1}", f"B{start_row + 1}", f"C{start_row + 1}"], TYPE_FILL)

    peak_count = trend_counter.most_common(1)[0][1]
    row_idx = start_row + 2
    for period, count in sorted(trend_counter.items()):
        summary.cell(row_idx, 1).value = period
        summary.cell(row_idx, 2).value = count
        summary.cell(row_idx, 3).value = count / total_valid
        summary.cell(row_idx, 3).number_format = "0.00%"
        summary.cell(row_idx, 3).font = RED_FONT
        for col_idx in range(1, 4):
            summary.cell(row_idx, col_idx).alignment = WRAP_TOP
        if count == peak_count:
            summary.cell(row_idx, 1).font = BASE_BOLD_FONT
            summary.cell(row_idx, 2).font = BASE_BOLD_FONT
            summary.cell(row_idx, 3).font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFF0000")
        else:
            summary.cell(row_idx, 1).font = BASE_FONT
            summary.cell(row_idx, 2).font = BASE_FONT
        row_idx += 1

    chart = LineChart()
    chart.style = 13
    chart.height = 7
    chart.width = 12.5
    chart.legend = None
    data = Reference(summary, min_col=2, min_row=start_row + 1, max_row=row_idx - 1)
    cats = Reference(summary, min_col=1, min_row=start_row + 2, max_row=row_idx - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    if chart.ser:
        chart.ser[0].marker.symbol = "circle"
        chart.ser[0].marker.size = 7
        chart.ser[0].graphicalProperties.line.solidFill = "FF5B9BD5"
        chart.ser[0].graphicalProperties.line.width = 28575
        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True
    summary.add_chart(chart, f"F{start_row + 1}")


def build_workbook(
    source_ws,
    product,
    metadata,
    rows,
    valid_rows,
    chain_counter,
    type_counter,
    subissue_counter,
    severity_counter,
    macro_counter,
    trend_counter,
    trend_granularity,
    output_path,
    cleaned_ws=None,
):
    wb = Workbook()
    wb.remove(wb.active)
    if cleaned_ws is not None:
        copy_source_sheet(cleaned_ws, wb, "CleanedComments")
    copy_source_sheet(source_ws, wb, "TaggedComments")
    summary = wb.create_sheet("Summary")

    total_valid = max(len(valid_rows), 1)
    fills = category_fill_map([name for name, _ in type_counter.most_common()])
    left_end = write_source_classification_block(summary, chain_counter, total_valid, fills)
    sorted_end = write_sorted_block(summary, subissue_counter, total_valid, fills)
    converged_end = write_converged_block(summary, type_counter, macro_counter, total_valid, fills)

    text, highlight_terms, bold_span = build_summary_text(
        product, metadata, rows, valid_rows, type_counter, subissue_counter, macro_counter, trend_counter, trend_granularity
    )
    summary.merge_cells("R1:Y20")
    summary["R1"] = build_rich_text(text, highlight_terms, bold_span)
    summary["R1"].alignment = WRAP_TOP

    chart_start = max(left_end, sorted_end, converged_end, 22)
    add_time_trend(summary, trend_counter, trend_granularity, len(valid_rows), chart_start)

    for col in ("A", "B", "C", "G", "H", "L", "M", "P", "R"):
        for cell in summary[col]:
            if cell.row > 2:
                cell.alignment = WRAP_TOP
                cell.font = BASE_FONT

    summary.freeze_panes = "A3"
    summary.column_dimensions["A"].width = 16
    summary.column_dimensions["B"].width = 24
    summary.column_dimensions["C"].width = 28
    summary.column_dimensions["D"].width = 10
    summary.column_dimensions["E"].width = 12
    summary.column_dimensions["G"].width = 16
    summary.column_dimensions["H"].width = 24
    summary.column_dimensions["I"].width = 10
    summary.column_dimensions["J"].width = 12
    summary.column_dimensions["L"].width = 16
    summary.column_dimensions["M"].width = 16
    summary.column_dimensions["N"].width = 10
    summary.column_dimensions["O"].width = 12
    summary.column_dimensions["P"].width = 18
    summary.column_dimensions["R"].width = 18
    summary.column_dimensions["S"].width = 18
    summary.column_dimensions["T"].width = 18
    summary.column_dimensions["U"].width = 18
    summary.column_dimensions["V"].width = 18
    summary.column_dimensions["W"].width = 18
    summary.column_dimensions["X"].width = 18
    summary.column_dimensions["Y"].width = 18
    apply_font_and_alignment(summary, min_row=2, max_col=25)

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Build a summary scaffold workbook from tagged comments.")
    parser.add_argument("workbook", help="Path to a workbook containing TaggedComments or CleanedComments")
    parser.add_argument("--sheet", help="Optional input sheet name")
    parser.add_argument("--product", help="Override product name")
    parser.add_argument("--output", help="Output xlsx path")
    args = parser.parse_args()

    workbook_path = Path(args.workbook).resolve()
    wb = load_workbook(workbook_path, data_only=True)
    metadata = read_metadata(wb)
    source_ws = find_input_sheet(wb, args.sheet)
    cleaned_ws = wb["CleanedComments"] if "CleanedComments" in wb.sheetnames and source_ws.title != "CleanedComments" else None
    rows = list(row_dicts(source_ws))
    (
        valid_rows,
        tagged_rows,
        chain_counter,
        type_counter,
        subissue_counter,
        severity_counter,
        macro_counter,
        trend_counter,
        trend_granularity,
    ) = summarize(rows)
    product = args.product or next((row.get("product_name") for row in rows if row.get("product_name")), "目标产品")
    output_path = (
        Path(args.output).resolve()
        if args.output
        else workbook_path.with_name(f"{product}_summary_scaffold.xlsx")
    )
    build_workbook(
        source_ws,
        product,
        metadata,
        rows,
        valid_rows,
        chain_counter,
        type_counter,
        subissue_counter,
        severity_counter,
        macro_counter,
        trend_counter,
        trend_granularity,
        output_path,
        cleaned_ws=cleaned_ws,
    )
    print(f"Created {output_path}")


if __name__ == "__main__":
    main()
