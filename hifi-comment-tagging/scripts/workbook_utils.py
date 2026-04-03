#!/usr/bin/env python3
from __future__ import annotations

import re
from collections import Counter
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook


HEADER_ALIASES = {
    "product_name": ["sku", "产品", "产品名", "product", "model", "型号", "机型"],
    "store": ["店铺", "store", "shop"],
    "country": ["国家", "country"],
    "raw_comment": ["原文", "英文评论", "comment", "review", "买家备注", "buyer note", "buyer remark"],
    "translated_comment": ["中文翻译", "翻译", "中文", "translation"],
    "return_reason": ["退货原因", "reason", "return reason"],
    "return_time": ["退货时间", "return time", "return date", "refund date"],
    "source_row_id": ["原始行号", "row id", "source row"],
    "order_id": ["订单号", "order id"],
    "level_1": ["一级分类", "问题分类：1级", "分类", "人工标签"],
    "level_2": ["二级分类", "问题分类：2级", "二级分类（人工）", "细分"],
    "level_3": ["三级问题点", "三级分类", "问题分类：3级"],
    "level_4": ["四级归因"],
}

PRODUCT_PATTERN = re.compile(
    r"(?:BOX\s*X\d+|LC\d{2,3}|GR\d{2,3}|ZD\d{1,3}|ZA\d{1,3}|ZP\d{1,3}|MC\d{2,3}|P\d{1,3}|Q\d{1,3}|BT\d+[A-Z]*|K\d+[A-Z]*)",
    re.IGNORECASE,
)
INVALID_MARKERS = {
    "",
    "na",
    "n/a",
    "none",
    "null",
    "同上",
    "sameasabove",
    "same as above",
    "未提供",
}
HIGH_RISK_PATTERNS = [
    "故障",
    "无法",
    "不工作",
    "失效",
    "保护",
    "静音",
    "断续",
    "噪音",
    "爆音",
    "电流声",
    "hdmi",
    "cec",
]
MEDIUM_RISK_PATTERNS = [
    "兼容",
    "不匹配",
    "音量",
    "增益",
    "失真",
    "破音",
]
USER_REASON_PATTERNS = ["不再需要", "买错", "下错单", "替代品", "改变主意", "需求变化"]
QUALITY_PATTERNS = ["质量", "故障", "品质", "噪音", "兼容", "保护", "静音", "电源"]


def load_workbook_safe(path: str | Path, data_only: bool = True):
    return load_workbook(Path(path), data_only=data_only)


def normalize_header(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value).strip().lower())


def normalize_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"\s+", " ", text)
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", text)


def display_text(value) -> str:
    return str(value).strip() if value is not None else ""


def is_invalid_feedback(value) -> bool:
    normalized = normalize_text(value)
    return normalized in INVALID_MARKERS or not normalized


def infer_products_from_text(text: str) -> List[str]:
    if not text:
        return []
    found = []
    for match in PRODUCT_PATTERN.finditer(text.upper()):
        token = re.sub(r"\s+", " ", match.group(0).strip()).upper()
        if token not in found:
            found.append(token)
    return found


def canonical_product_name(name: str) -> str:
    return re.sub(r"\s+", " ", name.strip()).upper()


def detect_header_row(ws, scan_rows: int = 6) -> Tuple[int, Dict[str, int]]:
    best_row = 1
    best_score = -1
    best_map: Dict[str, int] = {}
    for row_idx in range(1, min(scan_rows, ws.max_row) + 1):
        row_headers = {
            col_idx: normalize_header(ws.cell(row_idx, col_idx).value)
            for col_idx in range(1, ws.max_column + 1)
        }
        column_map: Dict[str, int] = {}
        score = 0
        for semantic, aliases in HEADER_ALIASES.items():
            normalized_aliases = {normalize_header(alias) for alias in aliases}
            for col_idx, header in row_headers.items():
                if header and header in normalized_aliases:
                    column_map[semantic] = col_idx
                    score += 2 if semantic.startswith("level_") else 1
                    break
        if score > best_score:
            best_score = score
            best_row = row_idx
            best_map = column_map
    return best_row, best_map


def infer_source_type(sheet_name: str, column_map: Dict[str, int]) -> str:
    lowered = sheet_name.lower()
    if "售后" in sheet_name or "after" in lowered:
        return "after_sales"
    if "return_reason" in column_map:
        return "return"
    if "退货" in sheet_name or "refund" in lowered or "return" in lowered:
        return "return"
    if "评论" in sheet_name or "review" in lowered:
        return "review"
    if any(key in column_map for key in ("level_1", "level_2", "level_3", "level_4")):
        return "review"
    return "unknown"


def sheet_profile(ws, workbook_path: str | Path) -> Dict[str, object]:
    header_row, column_map = detect_header_row(ws)
    sheet_products = infer_products_from_text(ws.title)
    workbook_products = infer_products_from_text(Path(workbook_path).name)
    return {
        "sheet": ws.title,
        "header_row": header_row,
        "column_map": column_map,
        "source_type": infer_source_type(ws.title, column_map),
        "sheet_products": sheet_products,
        "workbook_products": workbook_products,
        "candidate": is_candidate_sheet(column_map),
    }


def is_candidate_sheet(column_map: Dict[str, int]) -> bool:
    has_text = any(
        name in column_map for name in ("translated_comment", "raw_comment", "return_reason")
    )
    has_labels = any(name in column_map for name in ("level_1", "level_2", "level_3", "level_4"))
    return has_text or has_labels


def has_text_columns(column_map: Dict[str, int]) -> bool:
    return any(name in column_map for name in ("translated_comment", "raw_comment", "return_reason"))


def row_value(ws, row_idx: int, column_map: Dict[str, int], semantic: str) -> str:
    column_idx = column_map.get(semantic)
    if not column_idx:
        return ""
    return display_text(ws.cell(row_idx, column_idx).value)


def choose_comment_text(record: Dict[str, str], include_return_reason: bool = True) -> str:
    keys = ["translated_comment", "raw_comment"]
    if include_return_reason:
        keys.append("return_reason")
    for key in keys:
        value = record.get(key, "")
        if not is_invalid_feedback(value):
            return value.strip()
    return ""


def bool_from_value(value) -> bool:
    if isinstance(value, bool):
        return value
    normalized = normalize_text(value)
    return normalized in {"1", "true", "yes", "y", "是"}


def derive_severity(record: Dict[str, str]) -> str:
    combined = " ".join(
        [
            record.get("cleaned_comment", ""),
            record.get("level_1", ""),
            record.get("level_2", ""),
            record.get("level_3", ""),
            record.get("level_4", ""),
        ]
    )
    lowered = combined.lower()
    if any(pattern.lower() in lowered for pattern in HIGH_RISK_PATTERNS):
        return "high"
    if any(pattern.lower() in lowered for pattern in MEDIUM_RISK_PATTERNS):
        return "medium"
    return "low"


def derive_is_user_reason(record: Dict[str, str]) -> bool:
    combined = " ".join(
        [record.get("level_1", ""), record.get("level_2", ""), record.get("level_3", "")]
    )
    return any(pattern in combined for pattern in USER_REASON_PATTERNS)


def derive_is_quality_risk(record: Dict[str, str]) -> bool:
    combined = " ".join(
        [
            record.get("cleaned_comment", ""),
            record.get("level_1", ""),
            record.get("level_2", ""),
            record.get("level_3", ""),
        ]
    )
    return any(pattern in combined for pattern in QUALITY_PATTERNS)


def collect_product_candidates(path: str | Path) -> List[str]:
    workbook_path = Path(path)
    candidates = []
    seen = set()
    for product in infer_products_from_text(workbook_path.name):
        if product not in seen:
            seen.add(product)
            candidates.append(product)
    wb = load_workbook_safe(workbook_path, data_only=True)
    for ws in wb.worksheets:
        header_row, column_map = detect_header_row(ws)
        candidate_sheet = is_candidate_sheet(column_map)
        for product in infer_products_from_text(ws.title):
            if candidate_sheet and product not in seen:
                seen.add(product)
                candidates.append(product)
        product_col = column_map.get("product_name")
        if product_col and has_text_columns(column_map):
            counter = Counter()
            for row_idx in range(header_row + 1, min(ws.max_row, header_row + 200) + 1):
                product = display_text(ws.cell(row_idx, product_col).value).upper()
                for token in infer_products_from_text(product):
                    counter[token] += 1
            for product, _ in counter.most_common():
                if product not in seen:
                    seen.add(product)
                    candidates.append(product)
    return candidates


def sheet_matches_product(profile: Dict[str, object], target_product: str) -> bool:
    target = canonical_product_name(target_product)
    known = set(profile.get("sheet_products", [])) | set(profile.get("workbook_products", []))
    return target in known or target.replace(" ", "") in profile["sheet"].upper().replace(" ", "")


def collect_rows_for_product(path: str | Path, target_product: str, sheet_name: Optional[str] = None):
    workbook_path = Path(path)
    wb = load_workbook_safe(workbook_path, data_only=True)
    target = canonical_product_name(target_product)
    collected = []
    profiles = []

    for ws in wb.worksheets:
        profile = sheet_profile(ws, workbook_path)
        profiles.append(profile)
        if sheet_name and ws.title != sheet_name:
            continue
        if not profile["candidate"]:
            continue

        header_row = profile["header_row"]
        column_map = profile["column_map"]
        if not has_text_columns(column_map):
            continue
        product_col = column_map.get("product_name")
        sheet_is_match = sheet_matches_product(profile, target)

        for row_idx in range(header_row + 1, ws.max_row + 1):
            product_value = display_text(ws.cell(row_idx, product_col).value) if product_col else ""
            row_products = infer_products_from_text(product_value)
            row_is_match = target in row_products if row_products else False
            if product_col and not row_is_match:
                continue
            if not product_col and not sheet_is_match:
                continue

            record = {
                "record_id": f"{target}-{ws.title}-{row_idx}",
                "product_name": target,
                "source_type": profile["source_type"],
                "source_sheet": ws.title,
                "source_row": row_idx,
                "store": row_value(ws, row_idx, column_map, "store") or "未提供",
                "country": row_value(ws, row_idx, column_map, "country") or "未提供",
                "raw_comment": row_value(ws, row_idx, column_map, "raw_comment"),
                "translated_comment": row_value(ws, row_idx, column_map, "translated_comment"),
                "return_reason": row_value(ws, row_idx, column_map, "return_reason"),
                "return_time": row_value(ws, row_idx, column_map, "return_time"),
                "level_1": row_value(ws, row_idx, column_map, "level_1"),
                "level_2": row_value(ws, row_idx, column_map, "level_2"),
                "level_3": row_value(ws, row_idx, column_map, "level_3"),
                "level_4": row_value(ws, row_idx, column_map, "level_4"),
            }
            has_free_text = bool(
                normalize_text(record.get("translated_comment")) or normalize_text(record.get("raw_comment"))
            )
            is_return_feedback = profile["source_type"] == "return" and any(
                column_map.get(name) for name in ("raw_comment", "translated_comment")
            )
            record["cleaned_comment"] = choose_comment_text(
                record, include_return_reason=not is_return_feedback
            )
            record["is_valid_feedback"] = (
                has_free_text and not is_invalid_feedback(record["cleaned_comment"])
                if is_return_feedback
                else not is_invalid_feedback(record["cleaned_comment"])
            )
            record["severity"] = derive_severity(record)
            record["is_user_reason"] = derive_is_user_reason(record)
            record["is_quality_risk"] = derive_is_quality_risk(record)
            collected.append(record)
    return collected, profiles
