#!/usr/bin/env python3
import argparse
import json
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any, Optional
from urllib.parse import urlencode

import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Convert competitor run manifest into a multi-product Excel workbook.")
    parser.add_argument("manifest_json", help="Path to competitor_run_manifest.json")
    parser.add_argument("output_xlsx", nargs="?", help="Output Excel path")
    parser.add_argument("--cache-path", help="Optional translation cache path")
    parser.add_argument("--no-translate", action="store_true", help="Disable Chinese translation columns")
    return parser.parse_args()


def safe_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return ", ".join(str(item) for item in value)
    return str(value)


def normalize_text(text: str) -> str:
    return " ".join(text.split())


def parse_review_date(date_text: Optional[str]) -> str:
    if not date_text:
        return ""
    marker = " on "
    if marker not in date_text:
        return date_text
    return date_text.split(marker, 1)[1].strip()


def set_column_widths(ws, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def fit_image(image_path: Path, max_width: int = 120, max_height: int = 90) -> XLImage:
    img = XLImage(str(image_path))
    with PILImage.open(image_path) as opened:
        width, height = opened.size
    scale = min(max_width / width, max_height / height, 1)
    img.width = int(width * scale)
    img.height = int(height * scale)
    return img


def find_product_image_path(image_dir: Optional[Path], asin: str) -> Optional[Path]:
    if not image_dir or not image_dir.exists() or not asin:
        return None
    for extension in (".jpg", ".jpeg", ".png", ".webp"):
        candidate = image_dir / f"{asin}{extension}"
        if candidate.exists():
            return candidate
    return None


def load_translation_cache(cache_path: Path) -> dict[str, str]:
    if not cache_path.exists():
        return {}
    try:
        return json.loads(cache_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def save_translation_cache(cache_path: Path, cache: dict[str, str]) -> None:
    cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def translate_text(text: str) -> str:
    params = urlencode(
        {
            "client": "gtx",
            "sl": "auto",
            "tl": "zh-CN",
            "dt": "t",
            "q": text,
        }
    )
    response = requests.get(
        f"https://translate.googleapis.com/translate_a/single?{params}",
        headers={"User-Agent": "Mozilla/5.0"},
        timeout=20,
    )
    response.raise_for_status()
    payload = response.json()
    translated = "".join(part[0] for part in payload[0] if part and part[0])
    return translated or text


def translate_pending_texts(pending_texts: list[str], cache_path: Path, cache: dict[str, str]) -> None:
    if not pending_texts:
        return
    with ThreadPoolExecutor(max_workers=8) as executor:
        future_map = {executor.submit(translate_text, text): text for text in pending_texts}
        completed = 0
        for future in as_completed(future_map):
            text = future_map[future]
            try:
                cache[text] = future.result()
            except Exception:
                cache[text] = text
            completed += 1
            if completed % 20 == 0:
                save_translation_cache(cache_path, cache)
    save_translation_cache(cache_path, cache)


def enrich_reviews_with_translation(review_rows: list[dict[str, Any]], cache_path: Path) -> None:
    cache = load_translation_cache(cache_path)
    pending = []
    seen = set()
    for row in review_rows:
        for field in ("review_title", "review_body"):
            normalized = normalize_text(safe_text(row.get(field)))
            if normalized and normalized not in cache and normalized not in seen:
                seen.add(normalized)
                pending.append(normalized)
    translate_pending_texts(pending, cache_path, cache)
    for row in review_rows:
        row["review_title_zh"] = cache.get(normalize_text(safe_text(row.get("review_title"))), safe_text(row.get("review_title")))
        row["review_body_zh"] = cache.get(normalize_text(safe_text(row.get("review_body"))), safe_text(row.get("review_body")))


def load_manifest(manifest_path: Path) -> dict[str, Any]:
    return json.loads(manifest_path.read_text(encoding="utf-8"))


def collect_review_rows(manifest: dict[str, Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    product_rows = []
    review_rows = []

    for product in manifest.get("products", []):
        product_rows.append(product)
        if product.get("status") != "success":
            continue
        json_path = Path(product["jsonPath"])
        if not json_path.exists():
            continue
        payload = json.loads(json_path.read_text(encoding="utf-8"))
        for review in payload.get("reviews", []):
            review_rows.append(
                {
                    "scenario_label": product.get("scenarioLabel", ""),
                    "matched_keywords": product.get("matchedKeywords", []),
                    "product_title": product.get("title", ""),
                    "asin": product.get("asin", ""),
                    "product_url": product.get("productUrl", ""),
                    "author": review.get("author", ""),
                    "rating": review.get("rating"),
                    "rating_text": review.get("ratingText", ""),
                    "review_title": review.get("title", ""),
                    "review_body": review.get("body", ""),
                    "review_date": parse_review_date(review.get("dateText")),
                    "verified_purchase": review.get("verifiedPurchase", False),
                    "source_views": review.get("sourceViews", []),
                    "images": review.get("images", []),
                }
            )

    return product_rows, review_rows


def build_candidates_sheet(ws, manifest: dict[str, Any]) -> None:
    headers = [
        "场景编号",
        "场景名称",
        "是否保留",
        "关键词来源",
        "ASIN",
        "商品标题",
        "商品图片",
        "商品链接",
        "价格",
        "星级",
        "评论数",
        "品牌",
        "是否入选执行",
        "执行排序",
    ]
    ws.append(headers)
    style_header(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    set_column_widths(
        ws,
        {
            "A": 10,
            "B": 20,
            "C": 10,
            "D": 20,
            "E": 16,
            "F": 50,
            "G": 18,
            "H": 48,
            "I": 10,
            "J": 8,
            "K": 12,
            "L": 18,
            "M": 12,
            "N": 10,
        },
    )

    selected_by_asin = {product["asin"]: product for product in manifest.get("selectedProducts", [])}
    scenario_by_id = {scenario["id"]: scenario for scenario in manifest.get("scenarios", [])}
    image_dir = Path(manifest["_productImageDir"]) if manifest.get("_productImageDir") else None

    state_path = Path(manifest.get("statePath", ""))
    candidates = []
    if state_path.exists():
      state = json.loads(state_path.read_text(encoding="utf-8"))
      candidates = state.get("candidates", [])
    else:
      candidates = manifest.get("selectedProducts", [])

    for candidate in candidates:
        scenario = scenario_by_id.get(candidate.get("scenarioId"), {})
        selected = selected_by_asin.get(candidate.get("asin"))
        ws.append(
            [
                candidate.get("scenarioId", ""),
                candidate.get("scenarioLabel", ""),
                "是" if scenario.get("kept", candidate.get("kept", True)) else "否",
                safe_text(candidate.get("matchedKeywords")),
                candidate.get("asin", ""),
                candidate.get("title", ""),
                "",
                candidate.get("productUrl", ""),
                candidate.get("price") if candidate.get("price") is not None else "",
                candidate.get("ratingAverage") if candidate.get("ratingAverage") is not None else "",
                candidate.get("ratingCount", 0),
                candidate.get("brand", ""),
                "是" if selected else "否",
                selected.get("executionRank", "") if selected else "",
            ]
        )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP_TOP

    for row_index in range(2, ws.max_row + 1):
        asin = safe_text(ws[f"E{row_index}"].value)
        image_path = find_product_image_path(image_dir, asin)
        if not image_path:
            continue
        ws.add_image(fit_image(image_path), f"G{row_index}")
        ws.row_dimensions[row_index].height = 80


def build_review_sheet(ws, review_rows: list[dict[str, Any]], include_translation: bool) -> None:
    headers = [
        "场景名称",
        "关键词来源",
        "商品标题",
        "ASIN",
        "用户名称",
        "评分",
        "评分文本",
        "评论标题",
    ]
    if include_translation:
        headers.append("评论标题中文")
    headers.append("评论内容")
    if include_translation:
        headers.append("评论内容中文")
    headers.extend(
        [
            "评论时间",
            "是否认证购买",
            "来源视图",
            "图片数量",
            "图片路径",
        ]
    )

    max_images = max((len(row.get("images", [])) for row in review_rows), default=0)
    headers.extend([f"图片{i}" for i in range(1, max_images + 1)])

    ws.append(headers)
    style_header(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    widths = {
        "A": 18,
        "B": 18,
        "C": 36,
        "D": 16,
        "E": 16,
        "F": 8,
        "G": 16,
        "H": 26,
    }
    current_column = 9
    if include_translation:
        widths["I"] = 26
        current_column += 1
    widths[get_column_letter(current_column)] = 52
    current_column += 1
    if include_translation:
        widths[get_column_letter(current_column)] = 52
        current_column += 1
    for width in [14, 12, 20, 10, 48]:
        widths[get_column_letter(current_column)] = width
        current_column += 1
    for offset in range(max_images):
        widths[get_column_letter(current_column + offset)] = 18
    set_column_widths(ws, widths)
    image_start_column = current_column

    for row_data in review_rows:
        images = row_data.get("images", [])
        row = [
            row_data.get("scenario_label", ""),
            safe_text(row_data.get("matched_keywords")),
            row_data.get("product_title", ""),
            row_data.get("asin", ""),
            row_data.get("author", ""),
            row_data.get("rating", ""),
            row_data.get("rating_text", ""),
            row_data.get("review_title", ""),
        ]
        if include_translation:
            row.append(row_data.get("review_title_zh", ""))
        row.append(row_data.get("review_body", ""))
        if include_translation:
            row.append(row_data.get("review_body_zh", ""))
        row.extend(
            [
                row_data.get("review_date", ""),
                "是" if row_data.get("verified_purchase") else "否",
                safe_text(row_data.get("source_views")),
                len(images),
                "\n".join(image.get("localPath", "") for image in images if image.get("localPath")),
            ]
        )
        row.extend([""] * max_images)
        ws.append(row)
        current_row = ws.max_row
        row_has_image = False

        for cell in ws[current_row]:
            cell.alignment = WRAP_TOP

        for index, image in enumerate(images[:max_images], start=image_start_column):
            local_path = image.get("localPath")
            if not local_path:
                continue
            image_path = Path(local_path)
            if not image_path.exists():
                continue
            ws.add_image(fit_image(image_path), f"{get_column_letter(index)}{current_row}")
            row_has_image = True

        ws.row_dimensions[current_row].height = 80 if row_has_image else 60


def build_summary_sheet(ws, manifest: dict[str, Any]) -> None:
    headers = [
        "场景名称",
        "商品标题",
        "ASIN",
        "商品链接",
        "价格",
        "星级",
        "搜索页评论数",
        "实际抓取评论数",
        "带图评论数",
        "图片总数",
        "JSON 路径",
        "图片目录",
        "抓取状态",
    ]
    ws.append(headers)
    style_header(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    set_column_widths(
        ws,
        {
            "A": 18,
            "B": 36,
            "C": 16,
            "D": 48,
            "E": 10,
            "F": 8,
            "G": 12,
            "H": 12,
            "I": 12,
            "J": 10,
            "K": 40,
            "L": 40,
            "M": 12,
        },
    )

    for product in manifest.get("products", []):
        ws.append(
            [
                product.get("scenarioLabel", ""),
                product.get("title", ""),
                product.get("asin", ""),
                product.get("productUrl", ""),
                product.get("price", ""),
                product.get("ratingAverage", ""),
                product.get("ratingCount", 0),
                product.get("reviewCount", 0),
                product.get("imageReviewCount", 0),
                product.get("imageCount", 0),
                product.get("jsonPath", ""),
                product.get("mediaDir", ""),
                product.get("status", ""),
            ]
        )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP_TOP


def main() -> None:
    args = parse_args()
    manifest_path = Path(args.manifest_json).expanduser().resolve()
    output_path = (
        Path(args.output_xlsx).expanduser().resolve()
        if args.output_xlsx
        else manifest_path.with_suffix(".xlsx")
    )
    cache_path = (
        Path(args.cache_path).expanduser().resolve()
        if args.cache_path
        else output_path.with_name(f"{output_path.stem}_translation_cache.json")
    )
    product_image_dir = output_path.with_name(f"{output_path.stem}_product_images")

    manifest = load_manifest(manifest_path)
    manifest["_productImageDir"] = str(product_image_dir)
    product_rows, review_rows = collect_review_rows(manifest)
    include_translation = not args.no_translate

    if include_translation:
        enrich_reviews_with_translation(review_rows, cache_path)

    wb = Workbook()
    candidates_ws = wb.active
    candidates_ws.title = "场景与候选商品"
    build_candidates_sheet(candidates_ws, manifest)

    details_ws = wb.create_sheet("评论明细")
    build_review_sheet(details_ws, review_rows, include_translation)

    summary_ws = wb.create_sheet("抓取汇总")
    build_summary_sheet(summary_ws, manifest)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(
        json.dumps(
            {
                "outputPath": str(output_path),
                "candidateCount": len(json.loads(Path(manifest["statePath"]).read_text(encoding="utf-8")).get("candidates", []))
                if manifest.get("statePath") and Path(manifest["statePath"]).exists()
                else len(manifest.get("selectedProducts", [])),
                "productCount": len(product_rows),
                "reviewRowCount": len(review_rows),
                "translationCachePath": str(cache_path) if include_translation else None,
                "sheetNames": wb.sheetnames,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
