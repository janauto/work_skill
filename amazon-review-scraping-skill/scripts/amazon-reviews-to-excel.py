#!/usr/bin/env python3
import argparse
import json
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Any, Optional
from urllib.parse import urlencode

import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Convert Amazon review JSON into an Excel workbook with images.")
    parser.add_argument("input_json", help="Path to the Amazon review JSON file")
    parser.add_argument("output_xlsx", nargs="?", help="Output Excel file path")
    parser.add_argument("--cache-path", help="Optional translation cache JSON path")
    parser.add_argument("--no-translate", action="store_true", help="Disable Chinese translation columns")
    return parser.parse_args()


def parse_review_date(date_text: Optional[str]) -> str:
    if not date_text:
        return ""
    marker = " on "
    if marker not in date_text:
        return date_text
    raw_date = date_text.split(marker, 1)[1].strip()
    try:
        return datetime.strptime(raw_date, "%d %B %Y").strftime("%Y-%m-%d")
    except ValueError:
        return raw_date


def fit_image(image_path: Path, max_width: int = 120, max_height: int = 90) -> XLImage:
    img = XLImage(str(image_path))
    with PILImage.open(image_path) as opened:
        width, height = opened.size
    scale = min(max_width / width, max_height / height, 1)
    img.width = int(width * scale)
    img.height = int(height * scale)
    return img


def style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def set_column_widths(ws, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def safe_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return ", ".join(str(item) for item in value)
    return str(value)


def normalize_text(text: str) -> str:
    return " ".join(text.split())


def load_translation_cache(cache_path: Path) -> dict[str, str]:
    if not cache_path.exists():
        return {}
    try:
        return json.loads(cache_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def save_translation_cache(cache_path: Path, cache: dict[str, str]) -> None:
    cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def translate_text(text: str) -> str:
    params = urlencode(
        {
            "client": "gtx",
            "sl": "auto",
            "tl": "zh-CN",
            "dt": "t",
            "q": text,
        }
    )
    url = f"https://translate.googleapis.com/translate_a/single?{params}"
    response = requests.get(
      url,
      headers={"User-Agent": "Mozilla/5.0"},
      timeout=20,
    )
    response.raise_for_status()
    payload = response.json()
    translated = "".join(part[0] for part in payload[0] if part and part[0])
    return translated or text


def translate_pending_texts(pending_texts: list[str], cache_path: Path, cache: dict[str, str]) -> None:
    if not pending_texts:
        return

    with ThreadPoolExecutor(max_workers=8) as executor:
        future_map = {executor.submit(translate_text, text): text for text in pending_texts}
        completed = 0
        for future in as_completed(future_map):
            text = future_map[future]
            try:
                cache[text] = future.result()
            except Exception:
                cache[text] = text
            completed += 1
            if completed % 20 == 0:
                save_translation_cache(cache_path, cache)
    save_translation_cache(cache_path, cache)


def enrich_reviews_with_translation(reviews: list[dict[str, Any]], cache_path: Path) -> None:
    cache = load_translation_cache(cache_path)
    texts_to_translate = []
    seen = set()

    for review in reviews:
        for field in ("title", "body"):
            normalized = normalize_text(safe_text(review.get(field)))
            if normalized and normalized not in cache and normalized not in seen:
                seen.add(normalized)
                texts_to_translate.append(normalized)

    translate_pending_texts(texts_to_translate, cache_path, cache)

    for review in reviews:
        review["titleZh"] = cache.get(normalize_text(safe_text(review.get("title"))), safe_text(review.get("title")))
        review["bodyZh"] = cache.get(normalize_text(safe_text(review.get("body"))), safe_text(review.get("body")))


def build_review_sheet(ws, reviews: list[dict[str, Any]], max_images: int, include_translation: bool) -> None:
    headers = [
        "评论ID",
        "用户名称",
        "评分",
        "评分文本",
        "评论标题",
    ]
    if include_translation:
        headers.append("评论标题中文")
    headers.extend(["评论内容"])
    if include_translation:
        headers.append("评论内容中文")
    headers.extend(
        [
            "评论时间",
            "原始时间文本",
            "是否认证购买",
            "图片数量",
            "来源视图",
            "图片路径",
        ]
    )
    headers.extend([f"图片{i}" for i in range(1, max_images + 1)])

    ws.append(headers)
    style_header(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    widths = {
        "A": 18,
        "B": 18,
        "C": 8,
        "D": 18,
        "E": 34,
    }
    current_column = 6
    if include_translation:
        widths[get_column_letter(current_column)] = 34
        current_column += 1
    widths[get_column_letter(current_column)] = 60
    current_column += 1
    if include_translation:
        widths[get_column_letter(current_column)] = 60
        current_column += 1
    for width in [14, 32, 12, 10, 24, 50]:
        widths[get_column_letter(current_column)] = width
        current_column += 1
    for offset in range(max_images):
        widths[get_column_letter(current_column + offset)] = 18
    set_column_widths(ws, widths)

    image_column_start = current_column

    for review in reviews:
        images = review.get("images") or []
        row = [
            safe_text(review.get("reviewId")),
            safe_text(review.get("author")),
            review.get("rating") or "",
            safe_text(review.get("ratingText")),
            safe_text(review.get("title")),
        ]
        if include_translation:
            row.append(safe_text(review.get("titleZh")))
        row.append(safe_text(review.get("body")))
        if include_translation:
            row.append(safe_text(review.get("bodyZh")))
        row.extend(
            [
                parse_review_date(review.get("dateText")),
                safe_text(review.get("dateText")),
                "是" if review.get("verifiedPurchase") else "否",
                len(images),
                safe_text(review.get("sourceViews")),
                "\n".join(img.get("localPath", "") for img in images if img.get("localPath")),
            ]
        )
        row.extend([""] * max_images)
        ws.append(row)
        current_row = ws.max_row
        row_has_image = False

        for cell in ws[current_row]:
          cell.alignment = WRAP_TOP

        for index, image_info in enumerate(images[:max_images], start=image_column_start):
            local_path = image_info.get("localPath")
            if not local_path:
                continue
            image_path = Path(local_path)
            if not image_path.exists():
                continue
            row_has_image = True
            ws.add_image(fit_image(image_path), f"{get_column_letter(index)}{current_row}")

        ws.row_dimensions[current_row].height = 75 if not row_has_image else 80


def build_image_sheet(ws, reviews: list[dict[str, Any]], include_translation: bool) -> None:
    headers = [
        "评论ID",
        "用户名称",
        "评分",
        "评论标题",
    ]
    if include_translation:
        headers.append("评论标题中文")
    headers.extend(["评论时间", "图片序号", "图片路径", "图片"])

    ws.append(headers)
    style_header(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    widths = {
        "A": 18,
        "B": 18,
        "C": 8,
        "D": 34,
    }
    current_column = 5
    if include_translation:
        widths["E"] = 34
        current_column = 6
    for width in [14, 10, 50, 18]:
        widths[get_column_letter(current_column)] = width
        current_column += 1
    set_column_widths(ws, widths)

    image_column = "I" if include_translation else "H"

    for review in reviews:
        for image_info in review.get("images") or []:
            local_path = image_info.get("localPath")
            row = [
                safe_text(review.get("reviewId")),
                safe_text(review.get("author")),
                review.get("rating") or "",
                safe_text(review.get("title")),
            ]
            if include_translation:
                row.append(safe_text(review.get("titleZh")))
            row.extend(
                [
                    parse_review_date(review.get("dateText")),
                    image_info.get("imageIndex") or "",
                    safe_text(local_path),
                    "",
                ]
            )
            ws.append(row)
            current_row = ws.max_row
            for cell in ws[current_row]:
                cell.alignment = WRAP_TOP

            if local_path and Path(local_path).exists():
                ws.add_image(fit_image(Path(local_path)), f"{image_column}{current_row}")
                ws.row_dimensions[current_row].height = 80
            else:
                ws.row_dimensions[current_row].height = 40


def build_summary_sheet(ws, payload: dict[str, Any], reviews: list[dict[str, Any]], cache_path: Optional[Path]) -> None:
    ws["A1"] = "商品 ASIN"
    ws["B1"] = payload.get("asin", "")
    ws["A2"] = "商品链接"
    ws["B2"] = payload.get("productUrl", "")
    ws["A3"] = "抓取时间"
    ws["B3"] = payload.get("fetchedAt", "")
    ws["A4"] = "评论总数"
    ws["B4"] = len(reviews)
    ws["A5"] = "带图评论数"
    ws["B5"] = sum(1 for review in reviews if review.get("images"))
    ws["A6"] = "图片总数"
    ws["B6"] = sum(len(review.get("images") or []) for review in reviews)
    ws["A7"] = "图片目录"
    ws["B7"] = payload.get("mediaDir", "")
    ws["A8"] = "翻译缓存"
    ws["B8"] = str(cache_path) if cache_path else ""

    for cell in ws["A1:A8"]:
        cell[0].fill = HEADER_FILL
        cell[0].font = HEADER_FONT

    rating_counter = Counter(review.get("rating") for review in reviews if review.get("rating") is not None)
    start_row = 11
    ws[f"A{start_row}"] = "评分"
    ws[f"B{start_row}"] = "数量"
    ws[f"A{start_row}"].fill = HEADER_FILL
    ws[f"B{start_row}"].fill = HEADER_FILL
    ws[f"A{start_row}"].font = HEADER_FONT
    ws[f"B{start_row}"].font = HEADER_FONT

    row = start_row + 1
    for rating in sorted(rating_counter, reverse=True):
        ws[f"A{row}"] = rating
        ws[f"B{row}"] = rating_counter[rating]
        row += 1

    set_column_widths(ws, {"A": 18, "B": 90})
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.alignment = WRAP_TOP


def main() -> None:
    args = parse_args()
    input_path = Path(args.input_json).expanduser().resolve()
    output_path = (
        Path(args.output_xlsx).expanduser().resolve()
        if args.output_xlsx
        else input_path.with_suffix(".xlsx")
    )
    cache_path = (
        Path(args.cache_path).expanduser().resolve()
        if args.cache_path
        else input_path.with_name(f"{input_path.stem}_translation_cache.json")
    )

    with input_path.open("r", encoding="utf-8") as file:
        payload = json.load(file)

    reviews = payload.get("reviews") or []
    include_translation = not args.no_translate
    if include_translation:
        enrich_reviews_with_translation(reviews, cache_path)

    max_images = max((len(review.get("images") or []) for review in reviews), default=0)

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "统计"
    build_summary_sheet(summary_ws, payload, reviews, cache_path if include_translation else None)

    review_ws = wb.create_sheet("评论汇总")
    build_review_sheet(review_ws, reviews, max_images, include_translation)

    image_ws = wb.create_sheet("图片明细")
    build_image_sheet(image_ws, reviews, include_translation)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(
        json.dumps(
            {
                "outputPath": str(output_path),
                "reviewCount": len(reviews),
                "imageReviewCount": sum(1 for review in reviews if review.get("images")),
                "imageCount": sum(len(review.get("images") or []) for review in reviews),
                "translationCachePath": str(cache_path) if include_translation else None,
                "sheetNames": wb.sheetnames,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
