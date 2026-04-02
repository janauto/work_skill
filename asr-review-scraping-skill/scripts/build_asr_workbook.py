#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import math
import os
import re
import subprocess
import time
from pathlib import Path
from textwrap import wrap
from urllib.parse import urlparse

import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage, ImageDraw, ImageFont, ImageOps


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_DATASET_ROOT = SCRIPT_DIR.parent / "runs" / "default"
ROOT = DEFAULT_DATASET_ROOT
INDEX_PATH = ROOT / "thread_index.json"
OUTPUT_PATH = ROOT / "ASR_切换相关用户内容_打标准备.xlsx"
CACHE_PATH = ROOT / "translation_cache.json"
ZHIPU_URL = "https://open.bigmodel.cn/api/paas/v4/chat/completions"
IMAGE_DIR = ROOT / "downloaded_images"
PREVIEW_DIR = ROOT / "preview_images"
TMP_DIR = ROOT / "tmp"
CACHE_DATA_DIR = Path.home() / "Library/Caches/Google/Chrome/Default/Cache/Cache_Data"

IMAGE_EXCLUDE_PATTERNS = (
    "/data/avatars/",
    "forum%20resources",
    "cdn.jsdelivr.net/joypixels",
    "/reactions",
    "product_images/fav.png",
    "/favicon",
    "favicon.",
)


LABEL_TO_THEME = {
    "analog_switcher": "模拟切换器",
    "digital_switcher": "数字切换器",
    "switching_preamp": "切换前级",
    "avr_switching": "AVR切换功能",
    "audio_hub": "家庭音频中枢",
}


PRIMARY_THEME_BY_THREAD = {
    "2478": "模拟切换器",
    "29053": "数字切换器",
    "14043": "模拟切换器",
    "69760": "数字切换器",
    "11205": "家庭音频中枢",
    "23778": "模拟切换器",
    "41686": "家庭音频中枢",
    "15328": "模拟切换器",
    "46704": "家庭音频中枢",
    "32859": "数字切换器",
    "7654": "数字切换器",
    "64768": "数字切换器",
    "59602": "AVR切换功能",
    "51119": "家庭音频中枢",
    "34023": "模拟切换器",
    "1560": "数字切换器",
    "2338": "切换前级",
    "4746": "家庭音频中枢",
    "41367": "模拟切换器",
    "50438": "模拟切换器",
    "51156": "模拟切换器",
    "50455": "模拟切换器",
    "18413": "切换前级",
    "51070": "模拟切换器",
    "2335": "切换前级",
    "57198": "家庭音频中枢",
    "50342": "AVR切换功能",
    "41795": "模拟切换器",
    "50423": "切换前级",
}


PRODUCT_BY_THREAD = {
    "2478": "通用 1进2出切换器",
    "29053": "多路 Toslink 切换器",
    "14043": "带音量 A/B 切换器",
    "69760": "AES/EBU 多进一出切换器",
    "11205": "一体化音频中枢方案",
    "23778": "Parasound SCAMP 替代方案",
    "41686": "多信源功放接入方案",
    "15328": "ARX RS-1 XLR AB Switcher",
    "46704": "Audio Hub 方案",
    "32859": "DAC 自动输入切换",
    "7654": "自动光纤音频切换器",
    "64768": "AliExpress 自动 Toslink 切换器",
    "59602": "AVR + 有源分频方案",
    "51119": "Arylic BP50",
    "34023": "平衡 XLR 切换器",
    "1560": "ViewHD HDMI 3x1 Switcher",
    "2338": "Schiit Sys 硬件拆解",
    "4746": "Source-Control Hub 设想",
    "41367": "输入信号切换器（继电器 vs 无源）",
    "50438": "音箱切换盒选型",
    "51156": "MT-ViKI MT-431AV",
    "50455": "NEOHIPO ET30",
    "18413": "Passive Preamp",
    "51070": "RCA A/B 遥控切换器",
    "2335": "Schiit Sys",
    "57198": "RME ADI-2 + miniDSP SHD 切换方案",
    "50342": "AVR 作为前级/切换方案",
    "41795": "NJ&FX AUDIO PW-6",
    "50423": "双前级切换/共存方案",
}


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
RESAMPLE = PILImage.Resampling.LANCZOS if hasattr(PILImage, "Resampling") else PILImage.LANCZOS
IMAGE_SIGNATURES: list[tuple[bytes, str]] = [
    (b"\x89PNG\r\n\x1a\n", ".png"),
    (b"\xff\xd8\xff", ".jpg"),
    (b"GIF87a", ".gif"),
    (b"GIF89a", ".gif"),
]


def configure_paths(dataset_root: Path) -> None:
    global ROOT, INDEX_PATH, OUTPUT_PATH, CACHE_PATH, IMAGE_DIR, PREVIEW_DIR, TMP_DIR
    ROOT = dataset_root.resolve()
    INDEX_PATH = ROOT / "thread_index.json"
    OUTPUT_PATH = ROOT / "ASR_切换相关用户内容_打标准备.xlsx"
    CACHE_PATH = ROOT / "translation_cache.json"
    IMAGE_DIR = ROOT / "downloaded_images"
    PREVIEW_DIR = ROOT / "preview_images"
    TMP_DIR = ROOT / "tmp"


def parse_thread_id(local_file: str) -> str:
    return Path(local_file).name.split("_", 2)[1]


def resolve_local_file(local_file: str) -> Path:
    path = Path(local_file)
    if path.is_absolute():
        return path
    return ROOT / path


def read_index() -> list[dict]:
    return json.loads(INDEX_PATH.read_text(encoding="utf-8"))


def trim_to_posts(markdown: str) -> str:
    text = markdown.split("Markdown Content:\n", 1)[1] if "Markdown Content:\n" in markdown else markdown
    for marker in ("\n[You must log in or register to reply here.]", "\n### Similar threads", "\n## Similar threads"):
        if marker in text:
            text = text.split(marker, 1)[0]
    first_post = text.find("*   [#1](")
    if first_post != -1:
        heading = text.rfind("\n#### ", 0, first_post)
        if heading != -1:
            text = text[heading + 1 :]
    return text


def clean_line(line: str) -> str:
    line = line.strip()
    if not line:
        return ""
    line = re.sub(r"!\[[^\]]*\]\([^)]+\)", "", line)
    line = re.sub(r"\[\]\([^)]+\)", "", line)
    line = re.sub(r"\[([^\]]+)\]\(([^)]+)\)", lambda m: m.group(1) if not m.group(1).startswith("http") else m.group(2), line)
    line = re.sub(r"^\*\s+", "", line)
    line = re.sub(r"\s+", " ", line).strip()
    return line


def normalize_image_alt(alt: str, url: str) -> str:
    alt = re.sub(r"^Image\s+\d+:\s*", "", alt.strip(), flags=re.I)
    if alt:
        return alt
    parsed = urlparse(url)
    filename = Path(parsed.path).name
    return filename or parsed.netloc or url


def is_decorative_image(alt: str, url: str) -> bool:
    if any(pattern in url for pattern in IMAGE_EXCLUDE_PATTERNS):
        return True
    label = normalize_image_alt(alt, url).strip().lower()
    if not label:
        return True
    if label in {"like", "audio science review (asr) forum"}:
        return True
    if re.fullmatch(r"[:;=8xX][-()DPp/|]+", label):
        return True
    if re.fullmatch(r":[a-z0-9_+\-]+:", label):
        return True
    return False


def extract_image_items(line: str) -> list[dict[str, str]]:
    items: list[dict[str, str]] = []
    for alt, url in re.findall(r"!\[([^\]]*)\]\((https://[^)]+)\)", line):
        if is_decorative_image(alt, url):
            continue
        items.append({"url": url, "alt": normalize_image_alt(alt, url)})
    return items


def load_cache() -> dict[str, str]:
    if CACHE_PATH.exists():
        return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    return {}


def save_cache(cache: dict[str, str]) -> None:
    CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    CACHE_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def get_zhipu_api_key() -> str:
    for name in ("ZHIPUAI_API_KEY", "ZHIPU_API_KEY", "BIGMODEL_API_KEY"):
        value = os.environ.get(name, "").strip()
        if value:
            return value
    return ""


def chunk_text(text: str, size: int = 3500) -> list[str]:
    if len(text) <= size:
        return [text]
    parts: list[str] = []
    remaining = text
    while remaining:
        if len(remaining) <= size:
            parts.append(remaining)
            break
        split_at = remaining.rfind("\n", 0, size)
        if split_at < size * 0.5:
            split_at = remaining.rfind(". ", 0, size)
        if split_at < size * 0.5:
            split_at = size
        parts.append(remaining[:split_at].strip())
        remaining = remaining[split_at:].strip()
    return [p for p in parts if p]


def translate_text(text: str, cache: dict[str, str], api_key: str) -> str:
    text = text.strip()
    if not text:
        return ""
    if text in cache:
        return cache[text]
    if not api_key:
        return ""
    translated_parts: list[str] = []
    for part in chunk_text(text):
        payload = {
            "model": "glm-4-flash",
            "temperature": 0.2,
            "messages": [
                {
                    "role": "system",
                    "content": (
                        "你是专业论坛评论翻译助手。请把用户评论翻译成自然、准确的简体中文。"
                        "保留品牌、产品型号、接口名、缩写、URL、列表结构和原有换行。"
                        "不要总结，不要解释，不要补充。只输出翻译结果。"
                    ),
                },
                {"role": "user", "content": part},
            ],
        }
        response = requests.post(
            ZHIPU_URL,
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            },
            json=payload,
            timeout=120,
        )
        response.raise_for_status()
        data = response.json()
        translated_parts.append(
            data.get("choices", [{}])[0].get("message", {}).get("content", "").strip()
        )
    translated = "\n".join(p for p in translated_parts if p).strip()
    cache[text] = translated
    return translated


def parse_posts(markdown: str) -> list[dict]:
    lines = trim_to_posts(markdown).splitlines()
    posts: list[dict] = []
    current_author = ""
    current_role = ""
    pending_date = ""
    thread_starter_flag = False
    i = 0

    while i < len(lines):
        line = lines[i].rstrip("\n")
        stripped = line.strip()

        m = re.match(r"^#### \[([^\]]+)\]", stripped)
        if m:
            current_author = m.group(1).strip()
            pending_date = ""
            i += 1
            continue
        m = re.match(r"^#### (.+)$", stripped)
        if m and not stripped.startswith("#### Go to page"):
            current_author = m.group(1).strip()
            pending_date = ""
            i += 1
            continue
        m = re.match(r"^##### (.+)$", stripped)
        if m:
            current_role = m.group(1).strip()
            i += 1
            continue
        m = re.match(r"^\*\s+\[#(\d+)\]\(([^)]+)\)", stripped)
        if m:
            post_no = int(m.group(1))
            post_url = m.group(2).strip()
            i += 1
            raw_lines: list[str] = []
            quote_lines: list[str] = []
            body_lines: list[str] = []
            image_items: dict[str, dict[str, str]] = {}

            while i < len(lines):
                nxt = lines[i].rstrip("\n")
                nxt_stripped = nxt.strip()
                if re.match(r"^\*\s+\[#\d+\]\(", nxt_stripped):
                    break
                if nxt_stripped.startswith("#### ") or nxt_stripped.startswith("### Similar threads") or "[You must log in or register to reply here.]" in nxt_stripped:
                    break
                if nxt_stripped.startswith("Reactions:") or nxt_stripped.startswith("[You must log in"):
                    i += 1
                    continue
                for item in extract_image_items(nxt):
                    image_items[item["url"]] = item
                cleaned = clean_line(nxt)
                if cleaned:
                    if cleaned in {"OP", "I", "D", "M", "F", "C", "A", "H", "P", "S", "R", "W", "B", "J", "T"}:
                        i += 1
                        continue
                    raw_lines.append(cleaned)
                    if cleaned.startswith(">"):
                        quote_lines.append(cleaned.lstrip(">").strip())
                    elif cleaned not in ("OP", "D", "Thread Starter"):
                        body_lines.append(cleaned)
                i += 1

            posts.append(
                {
                    "post_no": post_no,
                    "post_url": post_url,
                    "author": current_author or "Unknown",
                    "author_role": current_role,
                    "post_date": pending_date,
                    "is_thread_starter": "是" if thread_starter_flag or post_no == 1 else "否",
                    "quote_text": "\n".join(quote_lines).strip(),
                    "post_text": "\n".join(body_lines).strip(),
                    "raw_text": "\n".join(raw_lines).strip(),
                    "image_items": list(image_items.values()),
                }
            )
            pending_date = ""
            thread_starter_flag = False
            continue

        m = re.match(r"^\*\s+\[([^\]#][^\]]*)\]\(([^)]+/post-\d+[^)]*)\)", stripped)
        if m:
            pending_date = m.group(1).strip()
            i += 1
            continue
        if "Thread Starter" in stripped:
            thread_starter_flag = True
            i += 1
            continue

        i += 1

    return posts


def safe_image_extension(content_type: str, url: str) -> str:
    content_type = (content_type or "").lower()
    if "png" in content_type:
        return ".png"
    if "webp" in content_type:
        return ".webp"
    if "gif" in content_type:
        return ".gif"
    if "bmp" in content_type:
        return ".bmp"
    if re.search(r"\.(png|jpg|jpeg|webp|gif|bmp)(?:$|\?)", url, flags=re.I):
        ext = re.search(r"\.(png|jpg|jpeg|webp|gif|bmp)(?:$|\?)", url, flags=re.I).group(1)
        return f".{ext.lower()}"
    return ".jpg"


def is_asr_attachment_url(url: str) -> bool:
    return "audiosciencereview.com/forum/" in url


def load_url_in_background_chrome(url: str, wait_seconds: int = 5) -> None:
    open_script = f'''
tell application "Google Chrome"
    set w to make new window
    set bounds of w to {{3200, 40, 3800, 640}}
    set URL of active tab of w to "{url}"
    delay {wait_seconds}
    close w
end tell
'''
    try:
        subprocess.run(
            ["osascript", "-e", open_script],
            check=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
    except Exception:
        pass


def parse_content_length(data: bytes) -> int:
    match = re.search(rb"content-length:(\d+)", data, flags=re.I)
    return int(match.group(1)) if match else 0


def parse_content_type(data: bytes) -> str:
    match = re.search(rb"content-type:([^\x00]+)", data, flags=re.I)
    return match.group(1).decode("utf-8", errors="ignore").strip().lower() if match else ""


def detect_image_start(data: bytes, url: str) -> tuple[int, str] | None:
    starts: list[tuple[int, str]] = []
    for signature, ext in IMAGE_SIGNATURES:
        idx = data.find(signature)
        if idx != -1:
            starts.append((idx, ext))
    riff_idx = data.find(b"RIFF")
    if riff_idx != -1 and data[riff_idx + 8 : riff_idx + 12] == b"WEBP":
        starts.append((riff_idx, ".webp"))
    if not starts:
        return None
    return min(starts, key=lambda item: item[0])


def extract_image_from_cache_file(cache_file: Path, url: str, target_stem: str, index: int) -> Path | None:
    data = cache_file.read_bytes()
    url_bytes = url.encode("utf-8")
    if url_bytes not in data:
        return None
    start_info = detect_image_start(data, url)
    if start_info is None:
        return None
    start, signature_ext = start_info
    content_length = parse_content_length(data)
    if not content_length:
        return None
    body = data[start : start + content_length]
    if len(body) < content_length:
        return None
    content_type = parse_content_type(data)
    ext = safe_image_extension(content_type, url) if content_type else signature_ext
    if content_type == "image/jpeg" and ext == ".jpeg":
        ext = ".jpg"
    target_path = IMAGE_DIR / f"{target_stem}_{index:02d}{ext}"
    target_path.write_bytes(body)
    try:
        with PILImage.open(target_path) as _:
            return target_path
    except Exception:
        target_path.unlink(missing_ok=True)
        return None


def find_image_in_chrome_cache(url: str, target_stem: str, index: int, timeout_seconds: int = 18) -> Path | None:
    if not CACHE_DATA_DIR.exists():
        return None
    start_time = time.time()
    lower_bound = start_time - 5
    seen: set[Path] = set()
    while time.time() - start_time < timeout_seconds:
        files = [
            p
            for p in CACHE_DATA_DIR.iterdir()
            if p.is_file() and p.stat().st_mtime >= lower_bound
        ]
        files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
        for cache_file in files[:80]:
            if cache_file in seen:
                continue
            seen.add(cache_file)
            extracted = extract_image_from_cache_file(cache_file, url, target_stem, index)
            if extracted is not None:
                return extracted
        time.sleep(0.75)
    return None


def find_existing_cache_hit(url: str, target_stem: str, index: int) -> Path | None:
    if not CACHE_DATA_DIR.exists():
        return None
    try:
        result = subprocess.run(
            ["rg", "-a", "-l", "-F", url, str(CACHE_DATA_DIR)],
            check=False,
            capture_output=True,
            text=True,
        )
    except FileNotFoundError:
        return None
    if result.returncode not in (0, 1):
        return None
    paths = [Path(line.strip()) for line in result.stdout.splitlines() if line.strip()]
    for cache_file in paths:
        extracted = extract_image_from_cache_file(cache_file, url, target_stem, index)
        if extracted is not None:
            return extracted
    return None


def fetch_asr_image_via_browser_cache(url: str, target_stem: str, index: int) -> Path | None:
    IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    for existing in IMAGE_DIR.glob(f"{target_stem}_{index:02d}.*"):
        existing.unlink(missing_ok=True)
    cached = find_existing_cache_hit(url, target_stem, index)
    if cached is not None:
        return cached
    load_url_in_background_chrome(url, wait_seconds=6)
    return find_image_in_chrome_cache(url, target_stem, index)


def download_image(url: str, target_stem: str, index: int) -> Path | None:
    IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    if is_asr_attachment_url(url):
        return fetch_asr_image_via_browser_cache(url, target_stem, index)
    headers = {"User-Agent": "Mozilla/5.0"}
    try:
        response = requests.get(url, headers=headers, timeout=30)
        response.raise_for_status()
        if not response.headers.get("Content-Type", "").lower().startswith("image/"):
            return None
        ext = safe_image_extension(response.headers.get("Content-Type", ""), url)
        path = IMAGE_DIR / f"{target_stem}_{index:02d}{ext}"
        if path.exists():
            return path
        path.write_bytes(response.content)
        return path
    except Exception:
        return None


def render_placeholder_tile(post_uid: str, index: int, alt: str, url: str) -> Path:
    PREVIEW_DIR.mkdir(parents=True, exist_ok=True)
    path = PREVIEW_DIR / f"{post_uid}_tile_{index:02d}.jpg"
    width, height = 320, 220
    canvas = PILImage.new("RGB", (width, height), "#F6F7F9")
    draw = ImageDraw.Draw(canvas)
    font = ImageFont.load_default()
    draw.rectangle((0, 0, width - 1, height - 1), outline="#C8CDD4", width=2)
    draw.rectangle((0, 0, width, 36), fill="#D9E7F5")
    draw.text((14, 12), "ASR附件预览受限", fill="#1F4E78", font=font)

    title = normalize_image_alt(alt, url)
    lines: list[str] = []
    for paragraph in filter(None, [title, urlparse(url).netloc]):
        lines.extend(wrap(paragraph, width=30)[:4])
        if len(lines) >= 7:
            break
    y = 54
    for line in lines[:7]:
        draw.text((14, y), line, fill="#222222", font=font)
        y += 18
    canvas.save(path, format="JPEG", quality=85)
    return path


def render_more_tile(post_uid: str, remaining: int) -> Path:
    PREVIEW_DIR.mkdir(parents=True, exist_ok=True)
    path = PREVIEW_DIR / f"{post_uid}_tile_more.jpg"
    canvas = PILImage.new("RGB", (320, 220), "#EEF1F4")
    draw = ImageDraw.Draw(canvas)
    font = ImageFont.load_default()
    draw.rectangle((0, 0, 319, 219), outline="#C8CDD4", width=2)
    draw.text((120, 92), f"+{remaining} more", fill="#4A5560", font=font)
    canvas.save(path, format="JPEG", quality=85)
    return path


def make_preview(post_uid: str, image_items: list[dict[str, str]]) -> tuple[list[str], str]:
    PREVIEW_DIR.mkdir(parents=True, exist_ok=True)
    downloaded_paths: list[Path] = []
    image_urls: list[str] = []
    for idx, item in enumerate(image_items, start=1):
        url = item["url"]
        image_urls.append(url)
        path = download_image(url, post_uid, idx)
        if path is None:
            path = render_placeholder_tile(post_uid, idx, item["alt"], url)
        downloaded_paths.append(path)

    if not downloaded_paths:
        return image_urls, ""

    thumbs: list[PILImage.Image] = []
    tiles = downloaded_paths[:8]
    if len(downloaded_paths) > 8:
        tiles = downloaded_paths[:7] + [render_more_tile(post_uid, len(downloaded_paths) - 7)]

    thumb_size = (300, 220)
    for path in tiles:
        try:
            img = PILImage.open(path).convert("RGB")
            thumbs.append(ImageOps.contain(img, thumb_size))
        except Exception:
            continue

    if not thumbs:
        return image_urls, ""

    cols = 2
    padding = 8
    rows = math.ceil(len(thumbs) / cols)
    canvas_w = cols * thumb_size[0] + (cols + 1) * padding
    canvas_h = rows * thumb_size[1] + (rows + 1) * padding
    canvas = PILImage.new("RGB", (canvas_w, canvas_h), "white")

    for idx, thumb in enumerate(thumbs):
        row = idx // cols
        col = idx % cols
        x = padding + col * (thumb_size[0] + padding)
        y = padding + row * (thumb_size[1] + padding)
        x += (thumb_size[0] - thumb.width) // 2
        y += (thumb_size[1] - thumb.height) // 2
        canvas.paste(thumb, (x, y))

    preview_path = PREVIEW_DIR / f"{post_uid}.jpg"
    canvas.save(preview_path, format="JPEG", quality=92)
    return image_urls, str(preview_path)


def write_sheet(ws, rows: list[list], widths: dict[int, int], freeze: str = "A2") -> None:
    for row in rows:
        ws.append(row)
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_TOP
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP_TOP if isinstance(cell.value, str) and len(cell.value) > 40 else TOP
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build an ASR tagging workbook with post text, translations, images, and previews.")
    parser.add_argument("--dataset-root", default=str(DEFAULT_DATASET_ROOT), help="Dataset directory containing thread_index.json and raw_threads.")
    parser.add_argument("--index-file", help="Override thread index path. Defaults to <dataset-root>/thread_index.json.")
    parser.add_argument("--workbook-path", help="Override workbook output path. Defaults to <dataset-root>/ASR_切换相关用户内容_打标准备.xlsx.")
    parser.add_argument("--cache-file", help="Override translation cache path. Defaults to <dataset-root>/translation_cache.json.")
    parser.add_argument("--image-dir", help="Override original image output dir. Defaults to <dataset-root>/downloaded_images.")
    parser.add_argument("--preview-dir", help="Override preview image output dir. Defaults to <dataset-root>/preview_images.")
    parser.add_argument("--chrome-cache-dir", help="Override Chrome cache dir used for ASR attachment extraction.")
    return parser.parse_args()


def main() -> None:
    global INDEX_PATH, OUTPUT_PATH, CACHE_PATH, IMAGE_DIR, PREVIEW_DIR, CACHE_DATA_DIR

    args = parse_args()
    configure_paths(Path(args.dataset_root).expanduser().resolve())
    if args.index_file:
        INDEX_PATH = Path(args.index_file).expanduser().resolve()
    if args.workbook_path:
        OUTPUT_PATH = Path(args.workbook_path).expanduser().resolve()
    if args.cache_file:
        CACHE_PATH = Path(args.cache_file).expanduser().resolve()
    if args.image_dir:
        IMAGE_DIR = Path(args.image_dir).expanduser().resolve()
    if args.preview_dir:
        PREVIEW_DIR = Path(args.preview_dir).expanduser().resolve()
    if args.chrome_cache_dir:
        CACHE_DATA_DIR = Path(args.chrome_cache_dir).expanduser().resolve()

    items = read_index()
    post_rows: list[list] = []
    cache = load_cache()
    api_key = get_zhipu_api_key()
    preview_refs: dict[str, str] = {}

    for item in items:
        thread_id = parse_thread_id(item["file"])
        primary_theme = PRIMARY_THEME_BY_THREAD.get(thread_id, "待判断")
        secondary_themes = "、".join(LABEL_TO_THEME[label] for label in item["labels"] if LABEL_TO_THEME.get(label) and LABEL_TO_THEME[label] != primary_theme)
        topic_labels = "、".join([x for x in [primary_theme, secondary_themes] if x])
        product_name = PRODUCT_BY_THREAD.get(thread_id, item["title"])
        local_file = resolve_local_file(item["file"])
        markdown = local_file.read_text(encoding="utf-8")
        posts = parse_posts(markdown)

        for post in posts:
            zh_text = translate_text(post["post_text"], cache, api_key)
            image_urls, preview_path = make_preview(f"{thread_id}-{post['post_no']}", post["image_items"])
            preview_refs[f"{thread_id}-{post['post_no']}"] = preview_path
            row = [
                f"{thread_id}-{post['post_no']}",
                primary_theme,
                secondary_themes,
                topic_labels,
                product_name,
                item["title"],
                item["url"],
                post["post_no"],
                post["author"],
                post["author_role"],
                post["post_date"],
                post["is_thread_starter"],
                post["quote_text"],
                post["post_text"],
                zh_text,
                post["raw_text"],
                len(post["image_items"]),
                "\n".join(image_urls),
                "",
                item["file"],
                "",
                "",
                "",
                "",
                "",
            ]
            post_rows.append(row)
            if len(cache) % 25 == 0:
                save_cache(cache)

    save_cache(cache)

    wb = Workbook()
    ws_posts = wb.active
    ws_posts.title = "打标总表"
    post_headers = [[
        "post_uid",
        "primary_theme",
        "secondary_themes",
        "topic_labels",
        "product_name",
        "thread_title",
        "thread_url",
        "post_no",
        "author",
        "author_role",
        "post_date",
        "is_thread_starter",
        "quote_text",
        "post_text",
        "中文翻译",
        "raw_text",
        "image_count",
        "image_urls",
        "评论图片预览",
        "local_file",
        "一级标签",
        "二级标签",
        "三级标签",
        "情绪倾向",
        "备注",
    ]]
    write_sheet(
        ws_posts,
        post_headers + post_rows,
        {
            1: 14,
            2: 16,
            3: 22,
            4: 22,
            5: 28,
            6: 42,
            7: 54,
            8: 8,
            9: 18,
            10: 18,
            11: 16,
            12: 10,
            13: 34,
            14: 62,
            15: 62,
            16: 62,
            17: 10,
            18: 40,
            19: 32,
            20: 38,
            21: 14,
            22: 14,
            23: 14,
            24: 12,
            25: 24,
        },
    )

    preview_col = 19
    for row_idx in range(2, ws_posts.max_row + 1):
        post_uid = ws_posts.cell(row=row_idx, column=1).value
        preview_path = preview_refs.get(post_uid, "")
        if not preview_path:
            continue
        path = Path(preview_path)
        if not path.exists():
            continue
        try:
            img = XLImage(str(path))
            display_max_w = 260
            if img.width > display_max_w:
                ratio = display_max_w / img.width
                img.width = int(img.width * ratio)
                img.height = int(img.height * ratio)
            ws_posts.add_image(img, f"{get_column_letter(preview_col)}{row_idx}")
            ws_posts.row_dimensions[row_idx].height = max(90, img.height * 0.75)
        except Exception:
            continue

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"saved: {OUTPUT_PATH}")
    print(f"posts: {len(post_rows)}")
    print(f"translated_rows: {sum(1 for row in post_rows if row[14])}")
    print(f"used_zhipu_api: {'yes' if api_key else 'no'}")


if __name__ == "__main__":
    main()
