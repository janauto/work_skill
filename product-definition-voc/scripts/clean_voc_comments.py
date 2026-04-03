#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from pathlib import Path

from openpyxl import Workbook

from voc_workbook_utils import collect_rows, normalize_key, normalize_text


OUTPUT_COLUMNS = [
    "record_id",
    "focus_name",
    "asin",
    "product_name",
    "product_image",
    "product_link",
    "thread_url",
    "source_primary_theme",
    "source_secondary_theme",
    "source_topic_labels",
    "scene_name",
    "keyword_source",
    "source_type",
    "source_sheet",
    "source_row",
    "store",
    "country",
    "rating",
    "rating_text",
    "review_date",
    "raw_title",
    "translated_title",
    "raw_comment",
    "translated_comment",
    "cleaned_comment",
    "image_count",
    "image_refs",
    "is_valid_feedback",
    "drop_reason",
]

GENERIC_FILLERS = {
    "good",
    "great",
    "nice",
    "ok",
    "okay",
    "fine",
    "works",
    "perfect",
    "bad",
    "poor",
    "loveit",
    "不错",
    "很好",
    "好",
    "可以",
    "满意",
    "一般",
    "垃圾",
    "很好用",
}

AD_SPAM_PATTERNS = {
    "coupon",
    "promo code",
    "discount code",
    "contact me",
    "whatsapp",
    "telegram",
    "free sample",
    "cashback",
    "返现",
    "优惠码",
    "加微信",
    "联系我",
    "广告",
    "推广",
}

SERVICE_PATTERNS = {
    "fast shipping",
    "shipping was fast",
    "delivery was fast",
    "arrived on time",
    "arrived early",
    "packaged well",
    "customer service",
    "seller was great",
    "物流很快",
    "配送很快",
    "包装完好",
    "客服很好",
    "卖家很好",
}

PRODUCT_SIGNAL_PATTERNS = {
    "sound",
    "audio",
    "switch",
    "volume",
    "button",
    "noise",
    "quality",
    "compatible",
    "setup",
    "speaker",
    "headphone",
    "tv",
    "turntable",
    "preamp",
    "fit",
    "function",
    "works with",
    "音质",
    "切换",
    "音量",
    "噪音",
    "兼容",
    "按钮",
    "安装",
    "扬声器",
    "耳机",
    "电视",
    "转盘",
    "前级",
    "功能",
}


def is_empty_or_na(text: str) -> bool:
    return normalize_text(text) in {"", "na", "n/a", "none", "null"}


def is_pure_emoji_or_punct(text: str) -> bool:
    return bool(str(text).strip()) and not normalize_key(text)


def is_low_information(text: str) -> bool:
    normalized = normalize_key(text)
    if not normalized:
        return True
    if normalized in GENERIC_FILLERS:
        return True
    if len(normalized) <= 4 and normalize_text(text) in GENERIC_FILLERS:
        return True
    words = [word for word in re.split(r"\s+", normalize_text(text)) if word]
    return len(words) <= 2 and normalized in GENERIC_FILLERS


def is_spam_or_ad(text: str) -> bool:
    lowered = normalize_text(text)
    return any(pattern in lowered for pattern in AD_SPAM_PATTERNS)


def is_service_only(text: str) -> bool:
    lowered = normalize_text(text)
    if not any(pattern in lowered for pattern in SERVICE_PATTERNS):
        return False
    return not any(pattern in lowered for pattern in PRODUCT_SIGNAL_PATTERNS)


def dedupe_rows(rows):
    kept = []
    dropped = []
    seen = set()

    for row in rows:
        key = (
            normalize_key(row.get("focus_name") or row.get("product_name") or row.get("scene_name")),
            normalize_key(row["cleaned_comment"]),
        )
        if key in seen:
            row["is_valid_feedback"] = "FALSE"
            row["drop_reason"] = "duplicate"
            dropped.append(row)
            continue
        seen.add(key)
        kept.append(row)
    return kept, dropped


def classify_drop_reason(text: str) -> str:
    if is_empty_or_na(text):
        return "empty_or_na"
    if is_pure_emoji_or_punct(text):
        return "emoji_or_punctuation_only"
    if is_low_information(text):
        return "low_information"
    if is_spam_or_ad(text):
        return "spam_or_ad"
    if is_service_only(text):
        return "logistics_or_seller_only"
    return ""


def clean_rows(rows):
    kept = []
    dropped = []
    for row in rows:
        row = dict(row)
        drop_reason = classify_drop_reason(row["cleaned_comment"])
        if drop_reason:
            row["is_valid_feedback"] = "FALSE"
            row["drop_reason"] = drop_reason
            dropped.append(row)
            continue
        row["is_valid_feedback"] = "TRUE"
        row["drop_reason"] = ""
        kept.append(row)
    kept, duplicate_rows = dedupe_rows(kept)
    dropped.extend(duplicate_rows)
    return kept, dropped


def write_sheet(ws, rows):
    ws.append(OUTPUT_COLUMNS)
    for row in rows:
        ws.append([row.get(column, "") for column in OUTPUT_COLUMNS])
    ws.freeze_panes = "A2"
    for column_cells in ws.columns:
        header = str(column_cells[0].value or "")
        letter = column_cells[0].column_letter
        if header in {"cleaned_comment", "raw_comment", "translated_comment", "image_refs"}:
            ws.column_dimensions[letter].width = 42
        elif "title" in header:
            ws.column_dimensions[letter].width = 28
        else:
            ws.column_dimensions[letter].width = 16


def main():
    parser = argparse.ArgumentParser(description="Clean VOC review rows for product-definition analysis.")
    parser.add_argument("workbook", help="Path to the xlsx workbook")
    parser.add_argument("--focus", help="Product, category, or scenario focus")
    parser.add_argument("--sheet", help="Optional explicit sheet name")
    parser.add_argument("--output", help="Output xlsx path")
    args = parser.parse_args()

    workbook_path = Path(args.workbook).resolve()
    rows, profiles = collect_rows(workbook_path, focus=args.focus, sheet_name=args.sheet)
    if not rows:
        raise SystemExit("No candidate rows matched the given workbook or focus.")

    kept_rows, dropped_rows = clean_rows(rows)
    output_path = (
        Path(args.output).resolve()
        if args.output
        else workbook_path.with_name("voc_cleaned_comments.xlsx")
    )

    wb = Workbook()
    clean_ws = wb.active
    clean_ws.title = "CleanedComments"
    write_sheet(clean_ws, kept_rows)

    dropped_ws = wb.create_sheet("DroppedRows")
    write_sheet(dropped_ws, dropped_rows)

    meta_ws = wb.create_sheet("Metadata")
    meta_ws.append(["field", "value"])
    metadata = {
        "source_workbook": str(workbook_path),
        "focus_name": args.focus or "",
        "matched_rows_before_cleaning": len(rows),
        "clean_rows": len(kept_rows),
        "dropped_rows": len(dropped_rows),
        "drop_reasons": dict(Counter(row["drop_reason"] for row in dropped_rows)),
        "profiled_sheets": [profile["sheet"] for profile in profiles],
    }
    for key, value in metadata.items():
        if isinstance(value, (dict, list)):
            value = json.dumps(value, ensure_ascii=False)
        meta_ws.append([key, value])

    wb.save(output_path)
    print(json.dumps({"output": str(output_path), **metadata}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
