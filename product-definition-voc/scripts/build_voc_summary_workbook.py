#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
from collections import Counter, defaultdict
from pathlib import Path
from tempfile import TemporaryDirectory
from urllib.parse import urlparse
from urllib.request import Request, urlopen, urlretrieve

from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage, ImageDraw, ImageFont


HEADER_FILL = PatternFill("solid", fgColor="FF4F81BD")
TYPE_FILL = PatternFill("solid", fgColor="FFD9EAF7")
SECTION_FILL = PatternFill("solid", fgColor="FFF2F2F2")
NEED_FILL = PatternFill("solid", fgColor="FFFFF2CC")
AHA_FILL = PatternFill("solid", fgColor="FFE2F0D9")
RISK_FILL = PatternFill("solid", fgColor="FFFDE9D9")
HEAT_FILL = PatternFill("solid", fgColor="FFEAF3FF")
HEADER_FONT = Font(bold=True, color="FFFFFFFF", name="Microsoft YaHei", size=11)
BASE_FONT = Font(name="Microsoft YaHei", size=11)
RED_FONT = Font(name="Microsoft YaHei", size=11, color="FFFF0000")
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")
VOC_CACHE_DIR = Path.home() / ".codex" / "cache" / "product-definition-voc"
VOC_CACHE_DIR.mkdir(parents=True, exist_ok=True)
PRODUCT_IMAGE_CACHE_DIR = VOC_CACHE_DIR / "product-images"
PRODUCT_IMAGE_CACHE_DIR.mkdir(parents=True, exist_ok=True)
MASLOW_CACHE_PATH = VOC_CACHE_DIR / "maslow-layer-cache.json"
PRODUCT_PAGE_TIMEOUT = 5
IMAGE_PREVIEW_MAX_SIZE = (420, 280)
MASLOW_MODEL_URL = "https://open.bigmodel.cn/api/paas/v4/chat/completions"
MASLOW_PROMPT_VERSION = "2026-04-03-v1"
HTTP_HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Accept-Language": "en-US,en;q=0.9",
}
DISPLAY_HEADER_MAP = {
    "功能/主题": "一级功能",
    "主要功能/主题": "一级功能",
}
FEATURE_DISPLAY_MAP = {
    "音量/控制": "音量",
    "音质/性能": "音质",
    "连接/兼容": "连接",
    "体积/部署": "体积",
    "做工/耐用": "做工",
    "价格/价值": "价格",
    "外观/标识": "外观",
    "前级/放大": "前级",
    "场景创新/组合玩法": "场景",
    "切换/路由": "切换",
    "基础功能": "基础",
}
SECONDARY_NEED_FALLBACK = {
    "音量": "调节更顺手",
    "音质": "切换后音质不掉",
    "连接": "接入更省心",
    "切换": "切换逻辑更直观",
    "前级": "前级链路更完整",
    "价格": "预算更可控",
    "做工": "用久了更放心",
    "外观": "操作识别更清楚",
    "体积": "摆放更轻松",
    "场景": "覆盖更多使用法",
    "基础": "把基础链路补齐",
}
NEGATIVE_EMOTION_LABELS = {"被噪音打断", "焦虑不确定", "失控难调", "失望落差", "愤怒拒绝", "烦躁麻烦"}
MASLOW_LAYER_ORDER = ["生理层", "安全层", "归属层", "尊重层", "自我实现层"]
MASLOW_LAYER_MAP = {
    "降摩擦": "生理层",
    "空间效率": "生理层",
    "连接确定性": "安全层",
    "控制确定性": "安全层",
    "音质稳定": "安全层",
    "稳定可靠": "安全层",
    "系统补足": "归属层",
    "场景适配": "归属层",
    "性价比安全感": "尊重层",
    "品质认同": "尊重层",
    "灵感启发": "自我实现层",
}
MASLOW_LAYER_DESCRIPTIONS = {
    "生理层": "先解决操作负担和空间占用，让系统先能顺手用起来。",
    "安全层": "先保证连接、控制、音质和可靠性，避免用户产生不确定感。",
    "归属层": "让设备真正融入既有系统和使用场景。",
    "尊重层": "让用户觉得这笔投入划算且体面。",
    "自我实现层": "激发新玩法和更高阶的使用想象。",
}
FEATURE_COLOR_MAP = {
    "连接": "FFEAF4E2",
    "切换": "FFFFF4CC",
    "音质": "FFDDEBF7",
    "音量": "FFFCE4D6",
    "前级": "FFE9E1F5",
    "价格": "FFF9DDE5",
    "做工": "FFE4ECF7",
    "外观": "FFFCEAD6",
    "体积": "FFE2F0D9",
    "场景": "FFF4E7D3",
    "基础": "FFF2F2F2",
}
FEATURE_FILL_MAP = {feature: PatternFill("solid", fgColor=color) for feature, color in FEATURE_COLOR_MAP.items()}
MASLOW_PYRAMID_COLOR_MAP = {
    "生理层": "#D8EFD3",
    "安全层": "#D9EAF7",
    "归属层": "#FCE5CD",
    "尊重层": "#EADCF8",
    "自我实现层": "#F4CCCC",
}
MASLOW_FALLBACK_HINTS = {
    "自我实现层": ("灵感", "启发", "玩法", "diy", "创作", "实验", "探索", "改装", "表达"),
    "尊重层": ("性价比", "预算", "值", "划算", "高级", "体面", "品质", "认同", "颜值"),
    "归属层": ("系统", "场景", "客厅", "桌面", "家庭", "融入", "搭配", "补足", "共存", "适配"),
    "安全层": ("稳定", "可靠", "确定", "兼容", "噪音", "底噪", "失真", "接口", "连接", "控制", "音质", "无声"),
    "生理层": ("顺手", "方便", "省事", "一键", "切换", "体积", "空间", "省步骤", "摩擦", "部署"),
}


def normalize_viewpoint_label(value) -> str:
    text = str(value or "").strip()
    return "疑似灌水" if text == "夸张/疑似灌水" else text


def feature_value(row) -> str:
    return normalize_feature_label(row.get("一级功能") or row.get("功能/主题"))


def row_dicts(ws):
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    for row_idx in range(2, ws.max_row + 1):
        row = {headers[col - 1]: ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)}
        if not any(value not in (None, "") for value in row.values()):
            continue
        yield row


def copy_sheet(source_ws, output_wb, title):
    copied = output_wb.create_sheet(title)
    for row_idx, row in enumerate(source_ws.iter_rows(), start=1):
        values = []
        for cell in row:
            value = cell.value
            if row_idx == 1 and isinstance(value, str):
                value = DISPLAY_HEADER_MAP.get(value, value)
            values.append(value)
        copied.append(values)
    copied.freeze_panes = "A2"
    for cell in copied[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_CENTER
    for row in copied.iter_rows(min_row=2):
        for cell in row:
            cell.font = BASE_FONT
            cell.alignment = WRAP_TOP
    return copied


def style_headers(ws, coords, fill):
    for coord in coords:
        ws[coord].fill = fill
        ws[coord].font = HEADER_FONT
        ws[coord].alignment = WRAP_CENTER


def feature_fill(feature: str) -> PatternFill:
    return FEATURE_FILL_MAP.get(str(feature or "").strip(), SECTION_FILL)


def fill_range(ws, row_idx: int, start_col: int, end_col: int, fill: PatternFill):
    for col_idx in range(start_col, end_col + 1):
        ws.cell(row_idx, col_idx).fill = fill


def get_bigmodel_api_key() -> str:
    for name in ("ZHIPUAI_API_KEY", "ZHIPU_API_KEY", "BIGMODEL_API_KEY"):
        value = os.environ.get(name, "").strip()
        if value:
            return value
    return ""


def load_json_cache(path: Path) -> dict:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_json_cache(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def collapse_text(value) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def semantic_fallback_layer(profile: dict) -> str:
    text = " ".join(
        [
            str(profile.get("底层需求", "")),
            str(profile.get("一级功能", "")),
            str(profile.get("二级功能需求", "")),
            str(profile.get("主要场景", "")),
            " ".join(profile.get("评论证据", []) or []),
        ]
    ).lower()
    for layer in ("自我实现层", "尊重层", "归属层", "安全层", "生理层"):
        if any(hint.lower() in text for hint in MASLOW_FALLBACK_HINTS[layer]):
            return layer
    return MASLOW_LAYER_MAP.get(str(profile.get("底层需求", "")).strip(), "安全层")


def trim_evidence(text: str, limit: int = 100) -> str:
    text = collapse_text(text)
    return text if len(text) <= limit else text[: limit - 1] + "…"


def build_need_profile(need: str, rows: list[dict]) -> dict:
    feature_counter = Counter(row["一级功能"] for row in rows)
    secondary_counter = Counter(row["二级功能需求"] for row in rows)
    scene_counter = Counter(row["场景标签"] for row in rows)
    ranked = sorted(rows, key=lambda row: (-int(row.get("排序分") or 0), -int(row.get("嘿哈分数") or 0)))
    evidences = []
    seen = set()
    for row in ranked:
        candidate = trim_evidence(row.get("translated_comment") or row.get("cleaned_comment") or row.get("raw_comment") or "")
        if len(candidate) < 18:
            continue
        normalized = candidate.lower()
        if normalized in seen:
            continue
        seen.add(normalized)
        evidences.append(candidate)
        if len(evidences) >= 3:
            break
    return {
        "底层需求": need,
        "数量": len(rows),
        "一级功能": feature_counter.most_common(1)[0][0],
        "二级功能需求": secondary_counter.most_common(1)[0][0],
        "主要场景": scene_counter.most_common(1)[0][0],
        "评论证据": evidences,
    }


def extract_json_blob(text: str) -> str:
    cleaned = str(text or "").strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned)
        cleaned = re.sub(r"\s*```$", "", cleaned)
    for opener, closer in (("[", "]"), ("{", "}")):
        start = cleaned.find(opener)
        end = cleaned.rfind(closer)
        if start != -1 and end != -1 and end > start:
            return cleaned[start : end + 1]
    return cleaned


def classify_maslow_with_bigmodel(profiles: list[dict], api_key: str) -> dict[str, dict]:
    if not profiles or not api_key:
        return {}
    system_prompt = (
        "你是产品定义阶段的 VOC 研究员。"
        "你要把每个需求簇归入且只能归入一个马斯洛层级：生理层、安全层、归属层、尊重层、自我实现层。"
        "判断必须基于评论语义，不允许只按关键词表面匹配。"
        "生理层=减少步骤、体力、空间和操作摩擦；"
        "安全层=稳定、可靠、确定、兼容、无风险；"
        "归属层=融入既有系统、生活场景和使用关系；"
        "尊重层=值不值、品质认同、体面感和被认可；"
        "自我实现层=探索新玩法、表达品位、创造性使用。"
        "只输出 JSON 数组，每项必须包含 底层需求、马斯洛层级、判断理由。"
    )
    payload = {
        "model": "glm-4-flash",
        "temperature": 0.1,
        "messages": [
            {"role": "system", "content": system_prompt},
            {
                "role": "user",
                "content": "请对下面这些需求簇做语义分层：\n" + json.dumps(profiles, ensure_ascii=False, indent=2),
            },
        ],
    }
    request = Request(
        MASLOW_MODEL_URL,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        method="POST",
    )
    with urlopen(request, timeout=120) as response:
        data = json.loads(response.read().decode("utf-8"))
    content = data.get("choices", [{}])[0].get("message", {}).get("content", "")
    blob = extract_json_blob(content)
    parsed = json.loads(blob) if blob else []
    results = {}
    for item in parsed if isinstance(parsed, list) else []:
        need = str(item.get("底层需求", "")).strip()
        layer = str(item.get("马斯洛层级", "")).strip()
        if need and layer in MASLOW_LAYER_ORDER:
            results[need] = {
                "layer": layer,
                "reason": collapse_text(item.get("判断理由", "")),
                "source": "bigmodel",
            }
    return results


def infer_maslow_layers(stats) -> dict[str, dict]:
    cache = load_json_cache(MASLOW_CACHE_PATH)
    api_key = get_bigmodel_api_key()
    ordered_needs = sorted(stats["need_rows"], key=lambda item: (-len(stats["need_rows"][item]), item))
    profiles = {need: build_need_profile(need, stats["need_rows"][need]) for need in ordered_needs}
    resolved = {}
    pending = []

    for need, profile in profiles.items():
        signature = hashlib.sha1(
            f"{MASLOW_PROMPT_VERSION}|{json.dumps(profile, ensure_ascii=False, sort_keys=True)}".encode("utf-8")
        ).hexdigest()
        cached = cache.get(signature)
        if cached and cached.get("layer") in MASLOW_LAYER_ORDER:
            resolved[need] = cached
        else:
            pending.append((need, signature, profile))

    if pending and api_key:
        try:
            llm_results = classify_maslow_with_bigmodel([profile for _need, _signature, profile in pending], api_key)
        except Exception:
            llm_results = {}
        for need, signature, profile in pending:
            if need in llm_results:
                cache[signature] = llm_results[need]
                resolved[need] = llm_results[need]

    for need, signature, profile in pending:
        if need in resolved:
            continue
        fallback_layer = semantic_fallback_layer(profile)
        resolved[need] = {
            "layer": fallback_layer,
            "reason": f"回退分层：需求语义更接近{fallback_layer}。",
            "source": "fallback",
        }
        cache[signature] = resolved[need]

    save_json_cache(MASLOW_CACHE_PATH, cache)
    return resolved


def load_chinese_font(size: int, bold: bool = False):
    candidates = [
        "/System/Library/Fonts/PingFang.ttc",
        "/System/Library/Fonts/Hiragino Sans GB.ttc",
        "/System/Library/Fonts/STHeiti Light.ttc",
        "/Library/Fonts/Arial Unicode.ttf",
    ]
    for candidate in candidates:
        path = Path(candidate)
        if path.exists():
            try:
                return ImageFont.truetype(str(path), size=size)
            except Exception:
                continue
    return ImageFont.load_default()


def create_maslow_pyramid_image(layer_counts: dict[str, int], total: int, temp_dir: Path) -> Path:
    width, height = 920, 560
    center_x = width // 2
    apex_y = 70
    base_y = 490
    half_base = 260
    image = PILImage.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(image)
    title_font = load_chinese_font(30, bold=True)
    label_font = load_chinese_font(24)
    stat_font = load_chinese_font(18)

    draw.text((40, 18), "马斯洛需求金字塔", fill="#1F1F1F", font=title_font)
    draw.text((40, 52), "按需求簇语义分层，优先反映用户真正想解决的问题。", fill="#666666", font=stat_font)

    def x_bounds(y: float) -> tuple[float, float]:
        ratio = (y - apex_y) / (base_y - apex_y)
        current_half = half_base * ratio
        return center_x - current_half, center_x + current_half

    top_to_bottom = list(reversed(MASLOW_LAYER_ORDER))
    for index, layer in enumerate(top_to_bottom):
        y_top = apex_y + (base_y - apex_y) * index / len(top_to_bottom)
        y_bottom = apex_y + (base_y - apex_y) * (index + 1) / len(top_to_bottom)
        left_top, right_top = x_bounds(y_top)
        left_bottom, right_bottom = x_bounds(y_bottom)
        polygon = [(left_top, y_top), (right_top, y_top), (right_bottom, y_bottom), (left_bottom, y_bottom)]
        draw.polygon(polygon, fill=MASLOW_PYRAMID_COLOR_MAP[layer], outline="#FFFFFF")
        count = layer_counts.get(layer, 0)
        percent = f"{(count / total):.1%}" if total else "0.0%"
        text = f"{layer}\n{count} 条 | {percent}"
        bbox = draw.multiline_textbbox((0, 0), text, font=label_font, spacing=6, align="center")
        text_x = center_x - (bbox[2] - bbox[0]) / 2
        text_y = (y_top + y_bottom) / 2 - (bbox[3] - bbox[1]) / 2
        draw.multiline_text((text_x, text_y), text, fill="#222222", font=label_font, spacing=6, align="center")

    image_path = temp_dir / "maslow_pyramid.png"
    image.save(image_path, format="PNG")
    return image_path


def set_column_widths(ws, widths):
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def apply_base_styles(ws, header_fill=HEADER_FILL):
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = HEADER_FONT
        cell.alignment = WRAP_CENTER
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BASE_FONT
            cell.alignment = WRAP_TOP


def apply_sheet_styles(ws, protected_rows: set[int]):
    for row in ws.iter_rows():
        for cell in row:
            if cell.row in protected_rows:
                continue
            if cell.value is not None:
                cell.font = BASE_FONT
                cell.alignment = WRAP_TOP


def build_rich_text(text: str, red_terms: list[str], bold_text: str) -> CellRichText:
    bold_start = text.find(bold_text)
    bold_end = bold_start + len(bold_text) if bold_start >= 0 else 0
    matches = []
    for term in sorted({term for term in red_terms if term}, key=len, reverse=True):
        for match in re.finditer(re.escape(term), text):
            start, end = match.span()
            if any(start < existing_end and end > existing_start for existing_start, existing_end in matches):
                continue
            matches.append((start, end))
    matches.sort()

    boundaries = {0, len(text), bold_start, bold_end}
    for start, end in matches:
        boundaries.add(start)
        boundaries.add(end)
    ordered = sorted(point for point in boundaries if point is not None)

    rich = CellRichText()
    for start, end in zip(ordered, ordered[1:]):
        if start == end:
            continue
        segment = text[start:end]
        if not segment:
            continue
        bold = bold_start <= start and end <= bold_end
        red = any(match_start <= start and end <= match_end for match_start, match_end in matches)
        if not bold and not red:
            rich.append(segment)
            continue
        font = InlineFont(rFont="Microsoft YaHei", sz=11, b=True if bold else None, color="FFFF0000" if red else None)
        rich.append(TextBlock(font, segment))
    return rich


def truncate_comment(text: str, length: int = 56) -> str:
    text = str(text or "").strip()
    return text if len(text) <= length else text[: length - 1] + "…"


def normalize_feature_label(feature) -> str:
    value = str(feature or "").strip()
    return FEATURE_DISPLAY_MAP.get(value, value or "基础")


def normalize_secondary_need(row) -> str:
    value = str(row.get("二级功能需求") or "").strip()
    if value:
        return value
    feature = feature_value(row)
    return SECONDARY_NEED_FALLBACK.get(feature, str(row.get("底层需求") or "").strip() or "把基础链路补齐")


def normalize_row(row):
    normalized = dict(row)
    feature = feature_value(row)
    normalized["一级功能"] = feature
    normalized["功能/主题"] = feature
    normalized["观点类型"] = normalize_viewpoint_label(row.get("观点类型"))
    normalized["情绪触发点"] = normalize_feature_label(row.get("情绪触发点") or feature)
    normalized["二级功能需求"] = normalize_secondary_need(normalized)
    return normalized


def source_comment_text(row) -> str:
    parts = [str(row.get("raw_title") or "").strip(), str(row.get("raw_comment") or "").strip()]
    source = "\n".join(part for part in parts if part)
    return source or str(row.get("cleaned_comment") or "").strip()


def build_aha_takeaway(row) -> str:
    product = str(row.get("product_name") or "该产品").strip()
    scene = str(row.get("场景标签") or "当前场景").strip()
    feature = feature_value(row)
    subneed = str(row.get("二级功能需求") or normalize_secondary_need(row)).strip()
    signal = str(row.get("决策信号") or "").strip()
    polarity = str(row.get("情绪极性") or "").strip()
    if polarity == "负向" or signal == "放弃采用":
        return f"{product} 在{scene}里必须把{subneed}做稳，否则用户不会继续采用。"
    if signal == "喜欢什么":
        return f"{product} 在{scene}里把{feature}做得更顺，用户最看重的是{subneed}。"
    if signal == "合理增配":
        return f"{product} 的方向被认可，但用户仍明确希望它把{subneed}做完整。"
    return f"{scene}用户需要一台把{subneed}做清楚的{feature}产品。"


def format_pairs(pairs, total: int, sep: str = "、") -> str:
    return sep.join(f"{name}{count / total:.2%}" for name, count in pairs) if pairs else "暂无"


def format_feature_secondary_pairs(pairs, total: int, sep: str = "、") -> str:
    if not pairs:
        return "暂无"
    return sep.join(f"{feature}-{secondary}{count / total:.2%}" for (feature, secondary), count in pairs)


def top_emotion_labels(stats, limit: int = 8) -> list[str]:
    labels = [label for label, _ in (stats["negative_emotions"] + stats["positive_emotions"]).most_common(limit)]
    labels = list(dict.fromkeys(labels))[:limit]
    return labels or ["满意认可"]


def write_heatmap_block(ws, start_row: int, title: str, row_header: str, row_names: list[str], labels: list[str], matrix_lookup):
    end_col = 1 + len(labels)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=end_col)
    ws.cell(start_row, 1).value = title
    style_headers(ws, [f"A{start_row}"], HEADER_FILL)
    ws.cell(start_row + 1, 1).value = row_header
    style_headers(ws, [f"A{start_row + 1}"], TYPE_FILL)
    for offset, label in enumerate(labels, start=2):
        column_letter = get_column_letter(offset)
        ws.cell(start_row + 1, offset).value = label
        style_headers(ws, [f"{column_letter}{start_row + 1}"], TYPE_FILL)

    current_row = start_row + 2
    for row_name in row_names:
        ws.cell(current_row, 1).value = row_name
        if row_header == "一级功能":
            ws.cell(current_row, 1).fill = feature_fill(row_name)
        for offset, label in enumerate(labels, start=2):
            ws.cell(current_row, offset).value = matrix_lookup[row_name].get(label, 0)
        current_row += 1
    apply_heatmap(ws, start_row + 2, current_row - 1, 2, end_col)
    return current_row - 1


def representative_phrase(row):
    feature = feature_value(row)
    if row.get("决策信号") == "喜欢什么":
        return f"用户看重 {feature or '该能力'} 带来的 {row.get('底层需求', '价值')}"
    if row.get("决策信号") == "放弃采用":
        return f"用户在 {feature or '该能力'} 上无法获得 {row.get('底层需求', '预期价值')}"
    if row.get("决策信号") == "合理增配":
        return f"用户希望进一步补足 {feature or '该能力'} 以强化 {row.get('底层需求', '价值')}"
    return truncate_comment(row.get("cleaned_comment", ""))


def parse_image_paths(value) -> list[Path]:
    refs = str(value or "")
    if not refs.strip():
        return []
    parts = re.split(r"\s*\|\s*|\n+", refs)
    paths = []
    for part in parts:
        path = Path(part.strip())
        if path.exists() and path.is_file():
            paths.append(path)
    return paths


def has_image_evidence(row) -> bool:
    return bool(
        split_image_refs(row.get("image_refs", ""))
        or split_image_refs(row.get("product_image", ""))
        or str(row.get("product_link", "")).strip()
    )


def image_evidence_score(row) -> int:
    if split_image_refs(row.get("image_refs", "")):
        return 2
    if split_image_refs(row.get("product_image", "")) or str(row.get("product_link", "")).strip():
        return 1
    return 0


def split_image_refs(value) -> list[str]:
    refs = str(value or "")
    if not refs.strip():
        return []
    return [part.strip() for part in re.split(r"\s*\|\s*|\n+", refs) if part.strip()]


def make_cache_key(value: str) -> str:
    return hashlib.sha1(value.encode("utf-8")).hexdigest()[:20]


def extract_asin(value: str) -> str:
    text = str(value or "")
    for pattern in (r"/dp/([A-Z0-9]{10})", r"/gp/product/([A-Z0-9]{10})", r"\b([A-Z0-9]{10})\b"):
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1).upper()
    return ""


def canonical_product_link(product_link: str) -> str:
    link = str(product_link or "").strip()
    if not link:
        return ""
    parsed = urlparse(link)
    asin = extract_asin(link)
    if asin and parsed.scheme and parsed.netloc:
        return f"{parsed.scheme}://{parsed.netloc}/dp/{asin}"
    return link


def resolve_image_ref(ref: str, temp_dir: Path) -> Path | None:
    if not ref:
        return None
    path = Path(ref)
    if path.exists() and path.is_file():
        return path
    if ref.startswith(("http://", "https://")):
        parsed = urlparse(ref)
        suffix = Path(parsed.path).suffix or ".jpg"
        filename = f"{make_cache_key(ref)}{suffix}"
        target = temp_dir / filename
        try:
            if not target.exists():
                urlretrieve(ref, target)
            return target if target.exists() else None
        except Exception:
            return None
    return None


def prepare_embed_image(source_path: Path, temp_dir: Path, embed_cache: dict[str, Path | None]) -> Path | None:
    source = Path(source_path)
    if not source.exists() or not source.is_file():
        return None
    cache_key = str(source.resolve())
    if cache_key in embed_cache:
        return embed_cache[cache_key]
    target = temp_dir / f"thumb_{make_cache_key(cache_key)}.jpg"
    try:
        with PILImage.open(source) as image:
            image = image.convert("RGB")
            image.thumbnail(IMAGE_PREVIEW_MAX_SIZE)
            image.save(target, format="JPEG", quality=82, optimize=True)
        embed_cache[cache_key] = target
    except Exception:
        embed_cache[cache_key] = source
    return embed_cache[cache_key]


def extract_main_image_from_product_link(product_link: str, temp_dir: Path, link_cache: dict[str, Path | None]) -> Path | None:
    if not product_link or not product_link.startswith(("http://", "https://")):
        return None
    if product_link in link_cache:
        return link_cache[product_link]
    cache_key = extract_asin(product_link) or make_cache_key(product_link)
    cache_path = PRODUCT_IMAGE_CACHE_DIR / f"{cache_key}.jpg"
    if cache_path.exists():
        link_cache[product_link] = cache_path
        return cache_path
    request_url = canonical_product_link(product_link)
    try:
        request = Request(request_url, headers=HTTP_HEADERS)
        with urlopen(request, timeout=PRODUCT_PAGE_TIMEOUT) as response:
            html = response.read().decode("utf-8", errors="ignore")
    except Exception:
        link_cache[product_link] = None
        return None

    patterns = [
        r'<meta[^>]+property=["\']og:image["\'][^>]+content=["\']([^"\']+)["\']',
        r'<meta[^>]+name=["\']twitter:image["\'][^>]+content=["\']([^"\']+)["\']',
        r'"large":"([^"]+)"',
        r'"mainUrl":"([^"]+)"',
    ]
    for pattern in patterns:
        match = re.search(pattern, html, re.IGNORECASE)
        if match:
            resolved = resolve_image_ref(match.group(1).replace("\\u0026", "&").replace("\\/", "/"), temp_dir)
            if resolved:
                try:
                    with PILImage.open(resolved) as image:
                        image = image.convert("RGB")
                        image.save(cache_path, format="JPEG", quality=88, optimize=True)
                    link_cache[product_link] = cache_path
                    return cache_path
                except Exception:
                    link_cache[product_link] = resolved
                    return resolved
    link_cache[product_link] = None
    return None


def resolve_image_evidence(
    user_refs,
    fallback_refs,
    product_link,
    temp_dir: Path,
    link_cache: dict[str, Path | None],
    embed_cache: dict[str, Path | None],
) -> tuple[list[Path], str]:
    resolved = [path for ref in split_image_refs(user_refs) if (path := resolve_image_ref(ref, temp_dir))]
    if resolved:
        prepared = [path for path in (prepare_embed_image(path, temp_dir, embed_cache) for path in resolved) if path]
        return prepared, str(user_refs or "")
    resolved = [path for ref in split_image_refs(fallback_refs) if (path := resolve_image_ref(ref, temp_dir))]
    if resolved:
        prepared = [path for path in (prepare_embed_image(path, temp_dir, embed_cache) for path in resolved) if path]
        return prepared, str(fallback_refs or "")
    product_image = extract_main_image_from_product_link(str(product_link or ""), temp_dir, link_cache)
    if not product_image:
        return [], ""
    prepared = prepare_embed_image(product_image, temp_dir, embed_cache)
    chosen_ref = canonical_product_link(str(product_link or "")) or str(product_link or "")
    return ([prepared] if prepared else []), f"产品图回退 | {chosen_ref}"


def add_images_to_row(ws, row_idx: int, start_col_letter: str, image_paths: list[Path], max_images: int = 2):
    if not image_paths:
        return
    anchors = [start_col_letter]
    if max_images > 1:
        anchors.extend(chr(ord(start_col_letter) + offset * 2) for offset in range(1, max_images))

    ws.row_dimensions[row_idx].height = 95
    for anchor_col, image_path in zip(anchors, image_paths[:max_images]):
        try:
            image = XLImage(str(image_path))
            max_width = 150
            max_height = 95
            scale = min(max_width / image.width, max_height / image.height, 1)
            image.width *= scale
            image.height *= scale
            ws.add_image(image, f"{anchor_col}{row_idx}")
        except Exception:
            continue


def pick_representative(rows, prefer_images: bool = False):
    def score(row):
        evidence_bonus = image_evidence_score(row) if prefer_images else 0
        return (evidence_bonus, int(row.get("嘿哈分数") or 0), int(row.get("排序分") or 0))

    return max(rows, key=score)


def apply_percent_style(ws, columns, start_row: int, end_row: int):
    for column in columns:
        for row_idx in range(start_row, end_row + 1):
            cell = ws[f"{column}{row_idx}"]
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00%"
                cell.font = RED_FONT


def apply_heatmap(ws, start_row: int, end_row: int, start_col: int, end_col: int):
    from openpyxl.utils import get_column_letter

    if end_row < start_row or end_col < start_col:
        return
    start = f"{get_column_letter(start_col)}{start_row}"
    end = f"{get_column_letter(end_col)}{end_row}"
    ws.conditional_formatting.add(
        f"{start}:{end}",
        ColorScaleRule(
            start_type="min",
            start_color="FFF2F2F2",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFFFE699",
            end_type="max",
            end_color="FFF8696B",
        ),
    )


def summarize(rows, raw_count):
    total = len(rows)
    feature_secondary_counter = Counter((row["一级功能"], row["二级功能需求"]) for row in rows)
    feature_scene_counter = Counter((row["一级功能"], row["场景标签"]) for row in rows)
    need_signal_counter = Counter((row["一级功能"], row["二级功能需求"]) for row in rows)
    need_counter = Counter(row["底层需求"] for row in rows)
    feature_counter = Counter(row["一级功能"] for row in rows)
    secondary_counter = Counter(row["二级功能需求"] for row in rows)
    scene_counter = Counter(row["场景标签"] for row in rows)
    emotion_feature_counter = defaultdict(Counter)
    emotion_scene_counter = defaultdict(Counter)
    feature_polarity_counter = defaultdict(Counter)
    scene_rows = defaultdict(list)
    need_rows = defaultdict(list)
    feature_rows = defaultdict(list)
    feature_secondary_rows = defaultdict(list)
    feature_scene_rows = defaultdict(list)
    aha_rows = []
    likes = Counter(row["二级功能需求"] for row in rows if row["决策信号"] == "喜欢什么")
    churn = Counter(row["二级功能需求"] for row in rows if row["决策信号"] == "放弃采用")
    requests = Counter(row["二级功能需求"] for row in rows if row["决策信号"] == "合理增配")
    negative_emotions = Counter(row["情绪标签"] for row in rows if row["情绪极性"] == "负向")
    positive_emotions = Counter(row["情绪标签"] for row in rows if row["情绪极性"] == "正向")
    aha_scene_counter = Counter()
    aha_secondary_counter = Counter()

    for row in rows:
        feature = row["一级功能"]
        secondary_need = row["二级功能需求"]
        scene = row["场景标签"]
        emotion_label = row.get("情绪标签", "观察确认")
        emotion_feature_counter[feature][emotion_label] += 1
        emotion_scene_counter[scene][emotion_label] += 1
        feature_polarity_counter[feature][row["情绪极性"]] += 1
        if row.get("情绪极性") == "负向" and row.get("情绪强度") == "强烈":
            feature_polarity_counter[feature]["强烈负向"] += 1
        scene_rows[scene].append(row)
        need_rows[row["底层需求"]].append(row)
        feature_rows[feature].append(row)
        feature_secondary_rows[(feature, secondary_need)].append(row)
        feature_scene_rows[(feature, scene)].append(row)
        if row.get("嘿哈时刻") == "是":
            aha_rows.append(row)
            aha_scene_counter[scene] += 1
            aha_secondary_counter[secondary_need] += 1

    aha_rows.sort(
        key=lambda row: (
            -int(row.get("嘿哈分数") or 0),
            row.get("嘿哈关联度") != "高",
            row.get("嘿哈独到性") != "高",
            -int(row.get("排序分") or 0),
        )
    )

    return {
        "total": total,
        "raw_count": raw_count,
        "feature_secondary_counter": feature_secondary_counter,
        "feature_scene_counter": feature_scene_counter,
        "need_signal_counter": need_signal_counter,
        "need_counter": need_counter,
        "feature_counter": feature_counter,
        "secondary_counter": secondary_counter,
        "scene_counter": scene_counter,
        "emotion_feature_counter": emotion_feature_counter,
        "emotion_scene_counter": emotion_scene_counter,
        "feature_polarity_counter": feature_polarity_counter,
        "scene_rows": scene_rows,
        "need_rows": need_rows,
        "feature_rows": feature_rows,
        "feature_secondary_rows": feature_secondary_rows,
        "feature_scene_rows": feature_scene_rows,
        "aha_rows": aha_rows,
        "likes": likes,
        "churn": churn,
        "requests": requests,
        "negative_emotions": negative_emotions,
        "positive_emotions": positive_emotions,
        "aha_scene_counter": aha_scene_counter,
        "aha_secondary_counter": aha_secondary_counter,
    }


def build_summary_text(focus_name: str, stats):
    total = max(stats["total"], 1)
    likes = stats["likes"].most_common(3)
    churn = stats["churn"].most_common(3)
    requests = stats["requests"].most_common(3)
    top_needs = stats["need_counter"].most_common(3)
    top_features = stats["feature_counter"].most_common(3)
    top_feature_secondary = stats["feature_secondary_counter"].most_common(3)
    top_scenes = stats["scene_counter"].most_common(3)
    top_aha_secondary = stats["aha_secondary_counter"].most_common(3)
    top_negative_emotions = stats["negative_emotions"].most_common(3)

    likes_text = format_pairs(likes, total)
    churn_text = format_pairs(churn, total)
    request_text = format_pairs(requests, total)
    need_text = format_pairs(top_needs, total)
    feature_text = format_pairs(top_features, total)
    feature_secondary_text = format_feature_secondary_pairs(top_feature_secondary, total)
    scene_text = format_pairs(top_scenes, total)
    aha_text = format_pairs(top_aha_secondary, total)
    emotion_text = format_pairs(top_negative_emotions, total)

    overview = (
        f"本次聚焦 {focus_name}，共清洗 {stats['raw_count']} 条评论，保留 {stats['total']} 条有效观点。"
        f"高频一级功能集中在 {feature_text}，一级功能下最聚焦的二级需求是 {feature_secondary_text}。"
        f"主要流失信号集中在 {churn_text}。"
    )

    lines = [
        "总结：",
        overview,
        "",
        f"1、用户喜欢什么\n用户稳定偏好主要收敛在 {likes_text}。这类评论更像是在说：只要把一级功能做清楚，并把对应的二级需求落到真实手感与链路细节上，用户就愿意长期采用。",
        "",
        f"2、为什么放弃采用\n主要流失信号集中在 {churn_text}。流失信号主要不是配置不够多，而是一级功能没把关键二级需求做稳。",
        "",
        f"3、一级功能与二级需求\n高频一级功能集中在 {feature_text}，对应最密集的二级需求是 {feature_secondary_text}。这说明产品定义阶段更适合先锁一级功能，再补清二级需求，而不是直接堆抽象卖点。",
        "",
        f"4、嘿哈时刻\n本次共识别 {len(stats['aha_rows'])} 条高价值嘿哈评论，最集中的启发方向是 {aha_text}。这些评论更接近产品定义灵感，而不是泛泛好评。",
        "",
        f"5、情绪热力图\n负向情绪最集中的标签是 {emotion_text}。这类情绪通常直接对应用户的拒买、弃用或不推荐行为，容错率最低。",
        "",
        f"6、场景机会与图片证据\n高频场景主要集中在 {scene_text}。场景卡和嘿哈板块中的图片证据可以直接还原真实使用环境，帮助判断需求是不是高频刚需。",
        "",
        f"7、底层需求归纳\n高频底层需求集中在 {need_text}。底层需求依然要看，但更适合放在一级功能和二级需求之后，用来解释为什么这些功能会成为刚需。",
        "",
        f"8、合理增配\n较集中的增配诉求集中在 {request_text}。更好的做法不是泛化堆功能，而是把这些诉求转成更明确的接口、调节逻辑、切换动作或反馈设计。",
    ]
    text = "\n".join(lines)
    red_terms = [
        focus_name,
        f"{stats['total']} 条",
        *(name for name, _ in top_features),
        *(f"{count / total:.2%}" for _, count in top_features),
        *(name for name, _ in likes),
        *(f"{count / total:.2%}" for _, count in likes),
        *(name for name, _ in churn),
        *(f"{count / total:.2%}" for _, count in churn),
        *(name for name, _ in top_needs),
        *(f"{count / total:.2%}" for _, count in top_needs),
        *(secondary for (_feature, secondary), _count in top_feature_secondary),
        *(name for name, _ in top_negative_emotions),
        *(f"{count / total:.2%}" for _, count in top_negative_emotions),
        *(name for name, _ in requests),
        *(f"{count / total:.2%}" for _, count in requests),
    ]
    return text, red_terms, overview


def write_need_clusters(wb, stats, temp_dir: Path):
    ws = wb.create_sheet("NeedClusters")
    total = max(stats["total"], 1)
    maslow_layers = stats["maslow_layers"]

    ws.merge_cells("A1:D1")
    ws["A1"] = "马斯洛层级分布"
    style_headers(ws, ["A1"], HEADER_FILL)
    for col, label in zip(("A", "B", "C", "D"), ("马斯洛层级", "数量", "占比", "层级说明")):
        ws[f"{col}2"] = label
    style_headers(ws, ["A2", "B2", "C2", "D2"], TYPE_FILL)

    layer_rows = defaultdict(list)
    for need, rows in stats["need_rows"].items():
        layer = maslow_layers.get(need, {}).get("layer", MASLOW_LAYER_MAP.get(need, "安全层"))
        layer_rows[layer].extend(rows)

    layer_counts = {}
    row_idx = 3
    for layer in MASLOW_LAYER_ORDER:
        rows = layer_rows.get(layer, [])
        count = len(rows)
        layer_counts[layer] = count
        ws.append([layer, count, count / total if total else 0, MASLOW_LAYER_DESCRIPTIONS[layer]])
        fill_range(ws, row_idx, 1, 4, PatternFill("solid", fgColor=FEATURE_COLOR_MAP.get("基础", "FFF2F2F2")))
        row_idx += 1

    pyramid_path = create_maslow_pyramid_image(layer_counts, total, temp_dir)
    pyramid = XLImage(str(pyramid_path))
    pyramid.width = 520
    pyramid.height = 320
    ws.add_image(pyramid, "J2")

    detail_start = 2 + len(MASLOW_LAYER_ORDER) + 2
    protected_rows = {1, 2, detail_start, detail_start + 1}
    ws.merge_cells(start_row=detail_start, start_column=1, end_row=detail_start, end_column=8)
    ws.cell(detail_start, 1).value = "需求簇明细：按马斯洛层级整理"
    style_headers(ws, [f"A{detail_start}"], HEADER_FILL)
    headers = ["马斯洛层级", "底层需求", "一级功能", "二级功能需求", "主要场景", "数量", "占比", "产品定义启发"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(detail_start + 1, idx).value = header
    style_headers(ws, [f"A{detail_start + 1}", f"B{detail_start + 1}", f"C{detail_start + 1}", f"D{detail_start + 1}", f"E{detail_start + 1}", f"F{detail_start + 1}", f"G{detail_start + 1}", f"H{detail_start + 1}"], TYPE_FILL)

    for need, rows in sorted(stats["need_rows"].items(), key=lambda item: (-len(item[1]), item[0])):
        feature_counter = Counter(row["一级功能"] for row in rows)
        secondary_counter = Counter(row["二级功能需求"] for row in rows)
        scene_counter = Counter(row["场景标签"] for row in rows)
        top_feature = feature_counter.most_common(1)[0][0]
        top_secondary = secondary_counter.most_common(1)[0][0]
        top_scene = scene_counter.most_common(1)[0][0]
        hint = f"产品定义可优先强化 {top_feature}，先把 {top_secondary} 做成对 {top_scene} 用户更明确的能力承诺。"
        layer = maslow_layers.get(need, {}).get("layer", MASLOW_LAYER_MAP.get(need, "安全层"))
        ws.append(
            [
                layer,
                need,
                top_feature,
                top_secondary,
                top_scene,
                len(rows),
                len(rows) / total,
                hint,
            ]
        )
        fill_range(ws, ws.max_row, 1, 8, feature_fill(top_feature))

    apply_percent_style(ws, ["C"], 3, 2 + len(MASLOW_LAYER_ORDER))
    apply_percent_style(ws, ["G"], detail_start + 2, ws.max_row)
    ws.freeze_panes = "A2"
    apply_sheet_styles(ws, protected_rows)
    set_column_widths(ws, {"A": 14, "B": 16, "C": 14, "D": 20, "E": 18, "F": 10, "G": 12, "H": 34, "J": 20, "K": 20, "L": 20, "M": 20})


def write_aha_sheet(wb, stats, temp_dir: Path, link_cache: dict[str, Path | None], embed_cache: dict[str, Path | None]):
    ws = wb.create_sheet("AhaMoments")
    headers = [
        "排名",
        "嘿哈分数",
        "关联度",
        "独到性",
        "商品名",
        "商品链接",
        "场景标签",
        "一级功能",
        "二级功能需求",
        "简明提炼",
        "源评论原文",
        "痛点机会",
        "产品定义启发",
        "图片证据1",
        "图片证据2",
    ]
    ws.append(headers)
    for index, row in enumerate(stats["aha_rows"], start=1):
        source_text = source_comment_text(row)
        ws.append(
            [
                index,
                row.get("嘿哈分数", ""),
                row.get("嘿哈关联度", ""),
                row.get("嘿哈独到性", ""),
                row.get("product_name", ""),
                row.get("product_link", ""),
                row.get("场景标签", ""),
                row.get("一级功能", ""),
                row.get("二级功能需求", ""),
                build_aha_takeaway(row),
                source_text,
                row.get("痛点机会", ""),
                row.get("产品定义启发", ""),
                "",
                "",
            ]
        )
        image_paths, _evidence_ref = resolve_image_evidence(
            row.get("image_refs", ""),
            row.get("product_image", ""),
            row.get("product_link", ""),
            temp_dir,
            link_cache,
            embed_cache,
        )
        add_images_to_row(ws, index + 1, "N", image_paths, max_images=2)
    apply_base_styles(ws, header_fill=AHA_FILL)
    set_column_widths(
        ws,
        {
            "A": 8,
            "B": 10,
            "C": 10,
            "D": 10,
            "E": 20,
            "F": 28,
            "G": 16,
            "H": 10,
            "I": 18,
            "J": 28,
            "K": 52,
            "L": 30,
            "M": 32,
            "N": 18,
            "O": 18,
        },
    )


def write_emotion_map(wb, stats):
    ws = wb.create_sheet("EmotionMap")
    total = max(stats["total"], 1)
    protected_rows = set()
    labels = top_emotion_labels(stats, 8)
    feature_rows = [feature for feature, _ in stats["feature_counter"].most_common()]
    scene_rows = [scene for scene, _ in stats["scene_counter"].most_common()]

    feature_end = write_heatmap_block(ws, 1, "功能 x 情绪标签热力图", "一级功能", feature_rows, labels, stats["emotion_feature_counter"])
    protected_rows.update({1, 2})
    scene_start = feature_end + 3
    scene_end = write_heatmap_block(ws, scene_start, "场景 x 情绪标签热力图", "场景标签", scene_rows, labels, stats["emotion_scene_counter"])
    protected_rows.update({scene_start, scene_start + 1})

    detail_start = scene_end + 3
    ws.merge_cells(start_row=detail_start, start_column=1, end_row=detail_start, end_column=6)
    ws.cell(detail_start, 1).value = "情绪标签明细：热区数据"
    style_headers(ws, [f"A{detail_start}"], HEADER_FILL)
    headers = ["一级功能", "情绪标签", "情绪极性", "数量", "占比", "热度说明"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(detail_start + 1, idx).value = header
    style_headers(ws, [f"A{detail_start + 1}", f"B{detail_start + 1}", f"C{detail_start + 1}", f"D{detail_start + 1}", f"E{detail_start + 1}", f"F{detail_start + 1}"], TYPE_FILL)

    label_polarity = {}
    for feature, counter in stats["emotion_feature_counter"].items():
        for label in counter:
            for scene_values in stats["scene_rows"].values():
                for row in scene_values:
                    if row["一级功能"] == feature and row.get("情绪标签") == label:
                        label_polarity[(feature, label)] = row.get("情绪极性", "")
                        break
                if (feature, label) in label_polarity:
                    break

    feature_emotion_pairs = []
    for feature, counter in stats["emotion_feature_counter"].items():
        for label, count in counter.items():
            feature_emotion_pairs.append((feature, label, count))

    sorted_pairs = sorted(feature_emotion_pairs, key=lambda item: (-item[2], item[0], item[1]))
    detail_row = detail_start + 2
    for feature, label, count in sorted_pairs:
        note = "高热负向" if label_polarity.get((feature, label)) == "负向" and count / total >= 0.03 else "重点观察"
        ws.append([feature, label, label_polarity.get((feature, label), ""), count, count / total, note])
        detail_row += 1

    protected_rows.update({detail_start, detail_start + 1})
    apply_percent_style(ws, ["E"], detail_start + 2, ws.max_row)
    ws.freeze_panes = "A2"
    apply_sheet_styles(ws, protected_rows)
    set_column_widths(ws, {"A": 18, "B": 18, "C": 10, "D": 10, "E": 12, "F": 14, "G": 14, "H": 14, "I": 14, "J": 14})


def write_scene_cards(wb, stats, temp_dir: Path, link_cache: dict[str, Path | None], embed_cache: dict[str, Path | None]):
    ws = wb.create_sheet("SceneCards")
    headers = [
        "场景标签",
        "商品名",
        "数量",
        "占比",
        "代表人物",
        "代表场地",
        "核心动机",
        "时空体验链路",
        "代表观点",
        "场景图片证据1",
        "场景图片证据2",
    ]
    ws.append(headers)
    total = max(stats["total"], 1)
    for scene, count in stats["scene_counter"].most_common():
        rows = stats["scene_rows"][scene]
        representative = pick_representative(rows, prefer_images=True)
        ws.append(
            [
                scene,
                representative.get("product_name", ""),
                count,
                count / total,
                representative.get("场景人物", ""),
                representative.get("场景地点", ""),
                representative.get("场景动机", ""),
                representative.get("场景链路", ""),
                representative_phrase(representative),
                "",
                "",
            ]
        )
        image_paths, _evidence_ref = resolve_image_evidence(
            representative.get("image_refs", ""),
            representative.get("product_image", ""),
            representative.get("product_link", ""),
            temp_dir,
            link_cache,
            embed_cache,
        )
        add_images_to_row(ws, ws.max_row, "J", image_paths, max_images=2)
    apply_base_styles(ws, header_fill=NEED_FILL)
    set_column_widths(ws, {"A": 18, "B": 18, "C": 10, "D": 12, "E": 14, "F": 16, "G": 18, "H": 36, "I": 28, "J": 18, "K": 18})
    apply_percent_style(ws, ["D"], 2, ws.max_row)


def write_summary_sheet(wb, stats, focus_name):
    summary = wb.create_sheet("Summary")
    total = max(stats["total"], 1)
    protected_rows = {1, 2}

    summary.merge_cells("A1:D1")
    summary["A1"] = "一级功能：先按话题看"
    style_headers(summary, ["A1"], HEADER_FILL)
    for col, label in zip(("A", "B", "C", "D"), ("一级功能", "二级功能需求", "数量", "占比")):
        summary[f"{col}2"] = label
    style_headers(summary, ["A2", "B2", "C2", "D2"], TYPE_FILL)

    row_idx = 3
    grouped = defaultdict(list)
    for (feature, secondary_need), count in stats["feature_secondary_counter"].items():
        grouped[feature].append(((feature, secondary_need), count))
    for feature, _count in stats["feature_counter"].most_common():
        items = grouped.get(feature, [])
        if not items:
            continue
        start = row_idx
        current_fill = feature_fill(feature)
        for index, ((feature_name, secondary_need), count) in enumerate(sorted(items, key=lambda item: (-item[1], item[0][1]))):
            summary.cell(row_idx, 1).value = feature_name if index == 0 else None
            summary.cell(row_idx, 2).value = secondary_need
            summary.cell(row_idx, 3).value = count
            summary.cell(row_idx, 4).value = count / total
            fill_range(summary, row_idx, 1, 4, current_fill)
            row_idx += 1
        if row_idx - 1 > start:
            summary.merge_cells(start_row=start, start_column=1, end_row=row_idx - 1, end_column=1)

    summary.merge_cells("G1:J1")
    summary["G1"] = "先看高频：一级功能筛选"
    style_headers(summary, ["G1"], HEADER_FILL)
    for col, label in zip(("G", "H", "I", "J"), ("一级功能", "二级功能需求", "数量", "占比")):
        summary[f"{col}2"] = label
    style_headers(summary, ["G2", "H2", "I2", "J2"], TYPE_FILL)
    high_row = 3
    for (feature, secondary_need), count in stats["feature_secondary_counter"].most_common(12):
        summary.cell(high_row, 7).value = feature
        summary.cell(high_row, 8).value = secondary_need
        summary.cell(high_row, 9).value = count
        summary.cell(high_row, 10).value = count / total
        fill_range(summary, high_row, 7, 10, feature_fill(feature))
        high_row += 1

    summary.merge_cells("L1:O1")
    summary["L1"] = "功能 x 场景排序"
    style_headers(summary, ["L1"], HEADER_FILL)
    for col, label in zip(("L", "M", "N", "O"), ("一级功能", "场景", "数量", "占比")):
        summary[f"{col}2"] = label
    style_headers(summary, ["L2", "M2", "N2", "O2"], TYPE_FILL)
    feature_scene_row = 3
    for (feature, scene), count in stats["feature_scene_counter"].most_common(12):
        summary.cell(feature_scene_row, 12).value = feature
        summary.cell(feature_scene_row, 13).value = scene
        summary.cell(feature_scene_row, 14).value = count
        summary.cell(feature_scene_row, 15).value = count / total
        fill_range(summary, feature_scene_row, 12, 15, feature_fill(feature))
        feature_scene_row += 1

    text, red_terms, overview = build_summary_text(focus_name, stats)
    summary.merge_cells("R1:Y24")
    summary["R1"] = build_rich_text(text, red_terms, overview)
    summary["R1"].alignment = WRAP_TOP

    block_start = max(row_idx, high_row, feature_scene_row, 27)

    summary.merge_cells(start_row=block_start, start_column=1, end_row=block_start, end_column=6)
    summary.cell(block_start, 1).value = "嘿哈摘要：高分评论"
    style_headers(summary, [f"A{block_start}"], HEADER_FILL)
    for col, label in zip(("A", "B", "C", "D", "E", "F"), ("排名", "商品名", "场景", "一级功能", "二级功能需求", "简明提炼")):
        summary[f"{col}{block_start + 1}"] = label
    style_headers(summary, [f"A{block_start + 1}", f"B{block_start + 1}", f"C{block_start + 1}", f"D{block_start + 1}", f"E{block_start + 1}", f"F{block_start + 1}"], TYPE_FILL)
    protected_rows.update({block_start, block_start + 1})
    for offset, row in enumerate(stats["aha_rows"][:5], start=block_start + 2):
        summary.cell(offset, 1).value = offset - (block_start + 1)
        summary.cell(offset, 2).value = row.get("product_name", "")
        summary.cell(offset, 3).value = row.get("场景标签", "")
        summary.cell(offset, 4).value = row.get("一级功能", "")
        summary.cell(offset, 5).value = row.get("二级功能需求", "")
        summary.cell(offset, 6).value = build_aha_takeaway(row)
        fill_range(summary, offset, 1, 6, feature_fill(row.get("一级功能", "")))

    scene_block_col = 8
    summary.merge_cells(start_row=block_start, start_column=scene_block_col, end_row=block_start, end_column=scene_block_col + 5)
    summary.cell(block_start, scene_block_col).value = "场景摘要：高频场景"
    style_headers(summary, [f"H{block_start}"], HEADER_FILL)
    for col, label in zip(("H", "I", "J", "K", "L", "M"), ("场景标签", "数量", "占比", "商品名", "人物", "代表观点")):
        summary[f"{col}{block_start + 1}"] = label
    style_headers(summary, [f"H{block_start + 1}", f"I{block_start + 1}", f"J{block_start + 1}", f"K{block_start + 1}", f"L{block_start + 1}", f"M{block_start + 1}"], TYPE_FILL)
    for offset, (scene, count) in enumerate(stats["scene_counter"].most_common(5), start=block_start + 2):
        representative = pick_representative(stats["scene_rows"][scene], prefer_images=True)
        summary.cell(offset, 8).value = scene
        summary.cell(offset, 9).value = count
        summary.cell(offset, 10).value = count / total
        summary.cell(offset, 11).value = representative.get("product_name", "")
        summary.cell(offset, 12).value = representative.get("场景人物", "")
        summary.cell(offset, 13).value = representative_phrase(representative)

    emotion_block_col = 15
    summary.merge_cells(start_row=block_start, start_column=emotion_block_col, end_row=block_start, end_column=emotion_block_col + 5)
    summary.cell(block_start, emotion_block_col).value = "情绪摘要：高热标签"
    style_headers(summary, [f"O{block_start}"], HEADER_FILL)
    for col, label in zip(("O", "P", "Q", "R", "S", "T"), ("情绪标签", "数量", "占比", "一级功能", "场景", "说明")):
        summary[f"{col}{block_start + 1}"] = label
    style_headers(summary, [f"O{block_start + 1}", f"P{block_start + 1}", f"Q{block_start + 1}", f"R{block_start + 1}", f"S{block_start + 1}", f"T{block_start + 1}"], TYPE_FILL)
    emotion_lookup_feature = defaultdict(Counter)
    emotion_lookup_scene = defaultdict(Counter)
    for feature, counter in stats["emotion_feature_counter"].items():
        for label, count in counter.items():
            emotion_lookup_feature[label][feature] += count
    for scene, counter in stats["emotion_scene_counter"].items():
        for label, count in counter.items():
            emotion_lookup_scene[label][scene] += count
    for offset, (label, count) in enumerate((stats["negative_emotions"] + stats["positive_emotions"]).most_common(5), start=block_start + 2):
        top_feature = emotion_lookup_feature[label].most_common(1)
        top_scene = emotion_lookup_scene[label].most_common(1)
        summary.cell(offset, 15).value = label
        summary.cell(offset, 16).value = count
        summary.cell(offset, 17).value = count / total
        summary.cell(offset, 18).value = top_feature[0][0] if top_feature else ""
        summary.cell(offset, 19).value = top_scene[0][0] if top_scene else ""
        summary.cell(offset, 20).value = "高热负向" if label in stats["negative_emotions"] else "稳定正向"

    chart_start = block_start + 9
    summary.merge_cells(start_row=chart_start, start_column=1, end_row=chart_start, end_column=11)
    summary.cell(chart_start, 1).value = "图表数据：Summary 可视化"
    style_headers(summary, [f"A{chart_start}"], HEADER_FILL)
    summary.cell(chart_start + 1, 1).value = "一级功能"
    summary.cell(chart_start + 1, 2).value = "占比"
    summary.cell(chart_start + 1, 4).value = "场景标签"
    summary.cell(chart_start + 1, 5).value = "占比"
    summary.cell(chart_start + 1, 7).value = "二级功能需求"
    summary.cell(chart_start + 1, 8).value = "数量"
    summary.cell(chart_start + 1, 10).value = "嘿哈场景"
    summary.cell(chart_start + 1, 11).value = "数量"
    style_headers(summary, [f"A{chart_start + 1}", f"B{chart_start + 1}", f"D{chart_start + 1}", f"E{chart_start + 1}", f"G{chart_start + 1}", f"H{chart_start + 1}", f"J{chart_start + 1}", f"K{chart_start + 1}"], TYPE_FILL)
    protected_rows.update({chart_start, chart_start + 1})

    top_features = stats["feature_counter"].most_common(8)
    for idx, (feature, count) in enumerate(top_features, start=chart_start + 2):
        summary.cell(idx, 1).value = feature
        summary.cell(idx, 2).value = count / total
    top_scenes = stats["scene_counter"].most_common(6)
    for idx, (scene, count) in enumerate(top_scenes, start=chart_start + 2):
        summary.cell(idx, 4).value = scene
        summary.cell(idx, 5).value = count / total
    top_secondary = stats["secondary_counter"].most_common(8)
    for idx, (secondary_need, count) in enumerate(top_secondary, start=chart_start + 2):
        summary.cell(idx, 7).value = secondary_need
        summary.cell(idx, 8).value = count
    for idx, (scene, count) in enumerate(stats["aha_scene_counter"].most_common(6), start=chart_start + 2):
        summary.cell(idx, 10).value = scene
        summary.cell(idx, 11).value = count

    if top_features:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "一级功能占比"
        chart.height = 6
        chart.width = 9
        data = Reference(summary, min_col=2, min_row=chart_start + 1, max_row=chart_start + 1 + len(top_features))
        cats = Reference(summary, min_col=1, min_row=chart_start + 2, max_row=chart_start + 1 + len(top_features))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        summary.add_chart(chart, f"M{chart_start}")

    if top_scenes:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "场景占比"
        chart.height = 6
        chart.width = 9
        data = Reference(summary, min_col=5, min_row=chart_start + 1, max_row=chart_start + 1 + len(top_scenes))
        cats = Reference(summary, min_col=4, min_row=chart_start + 2, max_row=chart_start + 1 + len(top_scenes))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        summary.add_chart(chart, f"U{chart_start}")

    if top_secondary:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "二级功能需求热度"
        chart.height = 6
        chart.width = 9
        data = Reference(summary, min_col=8, min_row=chart_start + 1, max_row=chart_start + 1 + len(top_secondary))
        cats = Reference(summary, min_col=7, min_row=chart_start + 2, max_row=chart_start + 1 + len(top_secondary))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        summary.add_chart(chart, f"M{chart_start + 14}")

    if stats["aha_scene_counter"]:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "嘿哈场景热度"
        chart.height = 6
        chart.width = 9
        data = Reference(summary, min_col=11, min_row=chart_start + 1, max_row=chart_start + 1 + len(stats["aha_scene_counter"]))
        cats = Reference(summary, min_col=10, min_row=chart_start + 2, max_row=chart_start + 1 + len(stats["aha_scene_counter"]))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        summary.add_chart(chart, f"U{chart_start + 14}")

    heat_start = chart_start + 14
    heat_labels = top_emotion_labels(stats, 6)
    feature_names = [feature for feature, _count in stats["feature_counter"].most_common(6)]
    write_heatmap_block(summary, heat_start, "功能 x 情绪标签热力图", "一级功能", feature_names, heat_labels, stats["emotion_feature_counter"])
    protected_rows.update({heat_start, heat_start + 1})

    apply_percent_style(summary, ["D", "J", "O", "Q"], 3, max(summary.max_row, 3))
    summary.freeze_panes = "A3"
    apply_sheet_styles(summary, protected_rows)
    set_column_widths(
        summary,
        {
            "A": 12,
            "B": 20,
            "C": 10,
            "D": 12,
            "G": 12,
            "H": 20,
            "I": 10,
            "J": 12,
            "L": 12,
            "M": 18,
            "N": 10,
            "O": 12,
            "R": 18,
            "S": 18,
            "T": 18,
            "U": 18,
            "V": 18,
            "W": 18,
            "X": 18,
            "Y": 18,
        },
    )


def main():
    parser = argparse.ArgumentParser(description="Build a styled VOC summary workbook from TaggedComments.")
    parser.add_argument("workbook", help="Path to workbook containing TaggedComments")
    parser.add_argument("--sheet", default="TaggedComments", help="Input sheet name")
    parser.add_argument("--focus", help="Override focus name")
    parser.add_argument("--output", help="Output xlsx path")
    args = parser.parse_args()

    workbook_path = Path(args.workbook).resolve()
    wb = load_workbook(workbook_path, data_only=True)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet '{args.sheet}' not found in {workbook_path}")
    tag_ws = wb[args.sheet]
    rows = [normalize_row(row) for row in row_dicts(tag_ws)]
    if not rows:
        raise SystemExit("No tagged rows found.")

    focus_name = args.focus or next((str(row.get("focus_name") or "") for row in rows if row.get("focus_name")), "VOC Focus")
    raw_count = len(rows)
    if "DroppedRows" in wb.sheetnames:
        raw_count += max(wb["DroppedRows"].max_row - 1, 0)

    stats = summarize(rows, raw_count)
    stats["maslow_layers"] = infer_maslow_layers(stats)
    output_path = Path(args.output).resolve() if args.output else workbook_path.with_name("voc_summary_workbook.xlsx")

    with TemporaryDirectory(prefix="voc-images-") as temp_dir_str:
        temp_dir = Path(temp_dir_str)
        link_cache: dict[str, Path | None] = {}
        embed_cache: dict[str, Path | None] = {}
        out_wb = Workbook()
        out_wb.remove(out_wb.active)
        if "CleanedComments" in wb.sheetnames:
            copy_sheet(wb["CleanedComments"], out_wb, "CleanedComments")
        copy_sheet(tag_ws, out_wb, "TaggedComments")
        write_need_clusters(out_wb, stats, temp_dir)
        write_aha_sheet(out_wb, stats, temp_dir, link_cache, embed_cache)
        write_emotion_map(out_wb, stats)
        write_scene_cards(out_wb, stats, temp_dir, link_cache, embed_cache)
        write_summary_sheet(out_wb, stats, focus_name)
        out_wb.save(output_path)
    print(output_path)


if __name__ == "__main__":
    main()
