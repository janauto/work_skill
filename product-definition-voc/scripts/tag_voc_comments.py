#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook

from voc_workbook_utils import normalize_key, normalize_text


TAG_COLUMNS = [
    "record_id",
    "focus_name",
    "asin",
    "product_name",
    "product_image",
    "product_link",
    "thread_url",
    "source_primary_theme",
    "source_secondary_theme",
    "source_topic_labels",
    "scene_name",
    "keyword_source",
    "source_type",
    "source_sheet",
    "source_row",
    "store",
    "country",
    "rating",
    "rating_text",
    "review_date",
    "raw_title",
    "translated_title",
    "raw_comment",
    "translated_comment",
    "cleaned_comment",
    "image_count",
    "image_refs",
    "观点类型",
    "一级功能",
    "二级功能需求",
    "底层需求",
    "决策信号",
    "情绪极性",
    "情绪强度",
    "情绪标签",
    "情绪触发点",
    "场景标签",
    "场景人物",
    "场景地点",
    "场景动机",
    "场景链路",
    "嘿哈时刻",
    "嘿哈分数",
    "嘿哈关联度",
    "嘿哈独到性",
    "痛点机会",
    "产品定义启发",
    "嘿哈分析",
    "排序分",
]

STRONG_POSITIVE = {
    "perfect",
    "finally",
    "exactly what i needed",
    "game changer",
    "brilliant",
    "love",
    "excellent",
    "amazing",
    "完美",
    "终于",
    "正是我需要的",
    "惊喜",
    "优秀",
    "太棒了",
    "非常适合",
}
POSITIVE = {
    "good",
    "great",
    "works well",
    "easy",
    "useful",
    "helpful",
    "clean",
    "clear",
    "solid",
    "不错",
    "很好",
    "简单",
    "方便",
    "有用",
    "好用",
    "清晰",
}
STRONG_NEGATIVE = {
    "garbage",
    "trash",
    "useless",
    "unusable",
    "broken",
    "dead",
    "垃圾",
    "无法使用",
    "没声音",
    "坏了",
    "失效",
    "退货",
}
NEGATIVE = {
    "bad",
    "noise",
    "hum",
    "buzz",
    "static",
    "crackle",
    "distortion",
    "delay",
    "issue",
    "problem",
    "annoying",
    "poor",
    "负担",
    "噪音",
    "电流声",
    "静电",
    "爆音",
    "问题",
    "差",
    "不稳定",
    "不兼容",
    "失真",
    "延迟",
}
CONVENIENCE_PATTERNS = {"easy", "convenient", "quick", "finally", "simple", "方便", "简单", "快速", "终于", "省事"}
NOISE_PATTERNS = {"noise", "hum", "buzz", "static", "crackle", "distortion", "噪音", "底噪", "电流声", "静电", "爆音", "失真"}
CONTROL_PATTERNS = {"volume", "mute", "button", "knob", "gain", "音量", "静音", "按钮", "旋钮", "增益"}
CONNECT_PATTERNS = {"connect", "compatible", "adapter", "rca", "xlr", "toslink", "spdif", "接口", "兼容", "连接", "接法"}
VALUE_PATTERNS = {"price", "worth", "cheap", "expensive", "价格", "值", "性价比", "便宜", "贵"}
SURPRISE_PATTERNS = {"finally", "exactly what i needed", "never thought", "surprised", "perfect", "正是我需要的", "终于", "没想到", "惊喜", "完美"}

DEFECT_PATTERNS = {
    "doesn't work",
    "doesnt work",
    "not working",
    "stopped working",
    "no sound",
    "one channel",
    "single channel",
    "failed after",
    "dead",
    "broken",
    "contact issue",
    "不起作用",
    "没声音",
    "停止工作",
    "坏了",
    "单声道",
    "声道异常",
    "接触不良",
}
SUGGESTION_PATTERNS = {
    "should",
    "wish",
    "would be better",
    "needs to",
    "need more",
    "i hope",
    "建议",
    "希望",
    "如果能",
    "应该",
    "最好",
    "需要增加",
}
COMPARE_PATTERNS = {
    "better than",
    "worse than",
    "compared to",
    "compared with",
    "instead of",
    "more than",
    "less than",
    "比",
    "相比",
    "替代",
    "取代",
    "对比",
}
EXPECTATION_PATTERNS = {
    "not as described",
    "i expected",
    "i thought",
    "misleading",
    "expected more",
    "description",
    "与描述不符",
    "预期",
    "以为",
    "误导",
    "没想到",
}
HYPE_PATTERNS = {
    "best ever",
    "life changing",
    "unbelievable",
    "must buy",
    "绝对必买",
    "神作",
    "无敌",
    "吹爆",
}

FEATURE_RULES = [
    ("音量", ("volume", "mute", "button", "knob", "remote", "gain", "音量", "静音", "按钮", "旋钮", "遥控", "增益")),
    ("音质", ("sound", "noise", "hum", "buzz", "static", "quality", "distortion", "latency", "delay", "音质", "噪音", "底噪", "失真", "延迟", "性能")),
    ("连接", ("connect", "compatible", "adapter", "hdmi", "usb", "rca", "xlr", "toslink", "spdif", "接口", "兼容", "连接", "接法")),
    ("体积", ("compact", "small", "desk", "desktop", "space", "setup", "install", "small footprint", "小巧", "紧凑", "桌面", "布置", "安装")),
    ("做工", ("build", "metal", "plastic", "sturdy", "durable", "broken", "quality", "做工", "耐用", "结构", "金属", "塑料", "质量")),
    ("价格", ("price", "cheap", "expensive", "value", "worth", "价格", "贵", "便宜", "值", "性价比")),
    ("外观", ("design", "look", "appearance", "label", "indicator", "led", "外观", "设计", "标识", "灯")),
    ("前级", ("preamp", "phono", "vinyl", "turntable", "amplifier", "唱机", "黑胶", "转盘", "前级", "放大")),
    ("场景", ("i use this to", "works with", "pair with", "route between", "用它来", "搭配", "组合", "玩法", "场景")),
    ("切换", ("switch", "selector", "route", "routing", "splitter", "input", "output", "切换", "切换器", "路由", "分配")),
]

NEED_RULES = [
    ("降摩擦", ("unplug", "switch between", "quickly", "finally", "one button", "不再反复插拔", "快速切换", "一步", "省事", "方便")),
    ("连接确定性", ("compatible", "works with", "connection", "stable signal", "兼容", "接法", "连接稳定", "信号稳定")),
    ("控制确定性", ("volume", "mute", "gain", "control", "predictable", "音量", "静音", "增益", "可控", "可预期")),
    ("音质稳定", ("sound quality", "clean sound", "no noise", "clear", "quality", "音质", "无噪音", "清晰", "性能")),
    ("稳定可靠", ("reliable", "durable", "last", "stopped working", "broken", "稳定", "可靠", "耐用", "故障", "失效")),
    ("空间效率", ("compact", "small", "hide", "desk", "space", "小巧", "紧凑", "省空间", "隐藏")),
    ("系统补足", ("turntable", "phono", "old receiver", "old system", "lack", "转盘", "黑胶", "旧系统", "补足", "补耳机口")),
    ("场景适配", ("multiple devices", "different outputs", "different sources", "versatile", "多设备", "多音源", "多场景", "灵活")),
    ("性价比安全感", ("price", "worth", "cheap", "expensive", "价格", "值", "性价比")),
    ("灵感启发", ("never thought", "surprised", "creative", "没想到", "惊喜", "灵感", "新玩法")),
    ("品质认同", ("premium", "quality feel", "build quality", "精致", "高级", "质感", "品质感")),
]

SCENE_RULES = [
    ("桌面双机", ("office", "work from home", "desktop", "computer", "pc", "mac", "zoom", "desk", "办公", "居家办公", "电脑", "桌面", "会议")),
    ("耳机音箱切换", ("headphone", "speaker", "headset", "earbud", "earphone", "耳机", "扬声器", "音箱")),
    ("家庭影音补足", ("tv", "television", "soundbar", "receiver", "home theater", "living room", "电视", "家庭影院", "客厅", "条形音箱", "光纤")),
    ("黑胶系统升级", ("turntable", "vinyl", "record", "phono", "old receiver", "唱机", "黑胶", "转盘", "旧系统")),
    ("录音创作", ("daw", "ableton", "protools", "synth", "keyboard", "guitar", "mixer", "broadcast", "录音", "直播", "创作", "合成器", "乐器")),
    ("游戏主机", ("xbox", "ps3", "ps4", "ps5", "wii", "console", "game", "游戏", "主机")),
    ("DIY项目", ("arduino", "raspberry", "robot", "relay", "project", "automation", "arduino", "树莓派", "项目", "自动化", "模块")),
    ("小空间部署", ("small", "compact", "apartment", "bedroom", "tiny desk", "小巧", "卧室", "小空间", "紧凑")),
]

SCENE_TEMPLATES = {
    "桌面双机": ("桌面多设备用户", "书桌或办公桌前", "在工作与娱乐设备之间快速切换", "在书桌前面对多设备输入输出，用户希望减少反复插拔和设置切换，快速进入正确的音频路径"),
    "耳机音箱切换": ("耳机与外放频繁切换的桌面用户", "电脑桌或监听位", "在私密聆听与外放之间无缝切换", "用户在同一套设备上交替使用耳机和音箱，希望一键完成切换且不破坏音质"),
    "家庭影音补足": ("家庭影音用户", "客厅电视或家庭影院旁", "给电视或旧影音系统补足缺失接口和控制能力", "用户在电视、条形音箱和耳机之间搭桥，希望补足现代电视缺少的独立音频控制"),
    "黑胶系统升级": ("黑胶或旧系统爱好者", "唱盘与功放之间", "让旧设备重新接入现代系统并保住听感", "用户围绕转盘、前级和功放重建链路，希望旧设备能够被现代系统继续使用"),
    "录音创作": ("创作者或录音用户", "录音台或工作站", "在监听、输出和表演设备之间灵活调度", "用户在创作链路里需要快速切换监听和输出，不想因路由成本打断创作"),
    "游戏主机": ("游戏或主机用户", "游戏桌或娱乐角", "在主机、显示器和声音输出之间保持顺滑体验", "用户需要让主机或娱乐设备接入现有声音链路，减少折腾和兼容问题"),
    "DIY项目": ("DIY 或自动化项目制作者", "工作台或项目箱体内", "在有限空间内实现稳定的信号控制", "用户把模块嵌入项目本体，希望结构薄、接线直观、功能确定"),
    "小空间部署": ("小空间布置用户", "卧室、出租屋或紧凑桌面", "以最低布置成本解决功能缺口", "用户在有限空间里搭系统，希望设备小、线少、上手快"),
    "通用补洞": ("通用系统补洞用户", "已有设备链路中", "补上一个原本缺失但高频需要的小能力", "用户并非追求复杂功能，而是希望用一个小设备填平体验断点"),
}

SOURCE_THEME_TO_FEATURE = [
    ("前级", ("切换前级", "前级", "preamp", "被动前级")),
    ("切换", ("模拟切换器", "数字切换器", "切换器", "切换功能", "ab switch", "switcher", "selector")),
    ("连接", ("家庭音频中枢", "audio hub", "hub", "中枢", "多信源", "接入方案", "avr")),
    ("音量", ("音量", "vu meter", "volume control", "gain")),
    ("音质", ("dac", "benchmark", "spdif", "toslink", "光纤", "数字音频")),
]

SOURCE_SECONDARY_HINTS = [
    ("模拟切换更直接", ("模拟切换器", "模拟", "rca", "xlr")),
    ("数字切换更稳定", ("数字切换器", "数字", "toslink", "spdif", "optical", "aes/ebu")),
    ("前级控制更完整", ("切换前级", "前级", "preamp", "被动前级")),
    ("多设备汇总更顺", ("家庭音频中枢", "audio hub", "hub", "中枢", "多信源")),
    ("旧系统接入更顺", ("avr", "receiver", "旧系统", "功放接入")),
]

SECONDARY_NEED_RULES = {
    "音量": [
        ("音量范围更大", ("louder", "more gain", "gain", "headroom", "output level", "更大音量", "音量不够", "推力", "增益")),
        ("电平匹配更准确", ("level match", "matching", "balance", "电平", "匹配", "平衡")),
        ("静音切换更干净", ("mute", "静音", "pop", "click", "爆音")),
        ("远程调节更方便", ("remote", "遥控", "couch", "sofa")),
    ],
    "音质": [
        ("底噪更低", ("noise", "hum", "buzz", "static", "底噪", "噪音", "电流声")),
        ("失真更少", ("distortion", "clipping", "失真", "削波")),
        ("声音更干净", ("clean", "clear", "transparent", "细节", "解析", "纯净")),
        ("延迟更低", ("latency", "delay", "延迟")),
    ],
    "连接": [
        ("接口覆盖更全", ("xlr", "rca", "toslink", "spdif", "optical", "hdmi", "usb", "6 or 8", "接口", "输入", "输出")),
        ("连接更稳定", ("stable", "intermittent", "drop", "dropout", "接触不良", "间歇", "稳定", "掉线")),
        ("兼容范围更广", ("compatible", "works with", "sample rate", "format", "兼容", "格式", "设备")),
        ("自动识别输入", ("auto", "automatic", "detect", "自动", "识别")),
    ],
    "切换": [
        ("一键切换更直接", ("a/b", "ab", "between", "switch between", "一键", "切换", "切来切去")),
        ("多路路由更清晰", ("2 out", "3 in", "4 in", "multi", "多路", "多输入", "多输出")),
        ("自动切换更聪明", ("auto", "automatic", "自动")),
        ("切换过程更安静", ("pop", "click", "mute", "爆音", "静音")),
    ],
    "前级": [
        ("补上前级控制", ("preamp", "volume", "前级", "音量控制")),
        ("黑胶链路补足前级", ("phono", "turntable", "vinyl", "唱机", "黑胶", "转盘")),
        ("推力或增益更充足", ("gain", "drive", "headphone", "amp", "推力", "增益")),
    ],
    "价格": [
        ("预算内更值", ("worth", "value", "cheap", "price", "性价比", "值", "便宜", "预算")),
        ("低成本替代更稳妥", ("alternative", "instead", "replace", "替代", "平替")),
    ],
    "做工": [
        ("结构更扎实", ("metal", "sturdy", "solid", "结构", "金属", "扎实")),
        ("寿命更长", ("durable", "last", "broken", "stopped working", "耐用", "寿命", "坏了")),
        ("手感更可靠", ("knob", "button", "switch feel", "旋钮", "按钮", "手感")),
    ],
    "外观": [
        ("状态反馈更清楚", ("indicator", "led", "label", "标识", "指示灯", "灯")),
        ("面板更直观", ("design", "look", "appearance", "设计", "外观", "面板")),
    ],
    "体积": [
        ("桌面占用更小", ("compact", "small", "desk", "desktop", "小巧", "紧凑", "桌面")),
        ("部署更省空间", ("setup", "install", "hide", "安装", "布置", "隐藏", "省空间")),
    ],
    "场景": [
        ("组合玩法更灵活", ("works with", "pair with", "route between", "搭配", "组合", "玩法")),
        ("一机多用更顺手", ("use this to", "versatile", "我用它来", "多用", "灵活")),
    ],
}


def row_dicts(ws):
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    for row_idx in range(2, ws.max_row + 1):
        row = {headers[col - 1]: ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)}
        if not any(value not in (None, "") for value in row.values()):
            continue
        yield row


def parse_rating(value) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        match = re.search(r"(\d+(?:\.\d+)?)", str(value))
        return float(match.group(1)) if match else 0.0


def contains_any(text: str, keywords) -> bool:
    return any(keyword in text for keyword in keywords)


def first_match(text: str, groups, default: str) -> str:
    for label, keywords in groups:
        if contains_any(text, keywords):
            return label
    return default


def source_theme_text(row) -> str:
    return normalize_text(
        " ".join(
            [
                str(row.get("source_primary_theme", "")),
                str(row.get("source_secondary_theme", "")),
                str(row.get("source_topic_labels", "")),
            ]
        )
    )


def classify_viewpoint(text: str, rating: float) -> str:
    if contains_any(text, HYPE_PATTERNS) and len(normalize_key(text)) < 50:
        return "疑似灌水"
    if contains_any(text, DEFECT_PATTERNS):
        return "缺陷复现"
    if contains_any(text, SUGGESTION_PATTERNS):
        return "建议"
    if contains_any(text, COMPARE_PATTERNS):
        return "对比"
    if contains_any(text, EXPECTATION_PATTERNS):
        return "期望落差"
    if rating and rating <= 2:
        return "抱怨"
    if contains_any(text, STRONG_NEGATIVE | NEGATIVE):
        return "抱怨"
    if rating and rating >= 4:
        return "优点"
    if contains_any(text, STRONG_POSITIVE | POSITIVE):
        return "优点"
    return "观点其他"


def classify_feature(text: str) -> str:
    return first_match(text, FEATURE_RULES, "基础")


def classify_feature_from_source(theme_text: str) -> str:
    return first_match(theme_text, SOURCE_THEME_TO_FEATURE, "")


def classify_secondary_need(text: str, feature: str, source_text: str = "") -> str:
    source_hint = first_match(source_text, SOURCE_SECONDARY_HINTS, "")
    if source_hint:
        source_feature_guard = {
            "模拟切换更直接": "切换",
            "数字切换更稳定": "切换",
            "前级控制更完整": "前级",
            "多设备汇总更顺": "连接",
            "旧系统接入更顺": "连接",
        }
        expected = source_feature_guard.get(source_hint)
        if expected in {"连接", "前级", "切换"} and feature == expected:
            return source_hint
        if source_hint == "多设备汇总更顺" and feature == "场景":
            return source_hint
    for label, keywords in SECONDARY_NEED_RULES.get(feature, []):
        if contains_any(text, keywords):
            return label
    fallback = {
        "音量": "调节更顺手",
        "音质": "切换后音质不掉",
        "连接": "接入更省心",
        "切换": "切换逻辑更直观",
        "前级": "前级链路更完整",
        "价格": "预算更可控",
        "做工": "用久了更放心",
        "外观": "操作识别更清楚",
        "体积": "摆放更轻松",
        "场景": "覆盖更多使用法",
        "基础": "把基础链路补齐",
    }
    return fallback.get(feature, "把基础链路补齐")


def classify_hidden_need(text: str, feature: str) -> str:
    need = first_match(text, NEED_RULES, "")
    if need:
        return need
    fallback = {
        "切换": "降摩擦",
        "连接": "连接确定性",
        "音量": "控制确定性",
        "音质": "音质稳定",
        "体积": "空间效率",
        "做工": "稳定可靠",
        "价格": "性价比安全感",
        "外观": "品质认同",
        "前级": "系统补足",
        "场景": "灵感启发",
    }
    return fallback.get(feature, "场景适配")


def classify_emotion(text: str, rating: float, viewpoint: str, feature: str):
    pos_score = sum(keyword in text for keyword in POSITIVE) + 2 * sum(keyword in text for keyword in STRONG_POSITIVE)
    neg_score = sum(keyword in text for keyword in NEGATIVE) + 2 * sum(keyword in text for keyword in STRONG_NEGATIVE)
    if rating >= 4:
        pos_score += 1
    elif 0 < rating <= 2:
        neg_score += 1

    if pos_score > neg_score:
        polarity = "正向"
    elif neg_score > pos_score:
        polarity = "负向"
    else:
        polarity = "中性"

    emphasis = text.count("!") + text.count("！")
    if rating in {1.0, 5.0} or emphasis >= 2 or contains_any(text, STRONG_POSITIVE | STRONG_NEGATIVE):
        intensity = "强烈"
    elif rating in {2.0, 4.0} or pos_score or neg_score:
        intensity = "中等"
    else:
        intensity = "轻微"

    if polarity == "正向":
        if contains_any(text, SURPRISE_PATTERNS):
            label = "惊喜超预期"
        elif contains_any(text, CONVENIENCE_PATTERNS):
            label = "顺手便利"
        elif feature in {"连接", "做工"}:
            label = "放心省心"
        elif feature == "价格" or contains_any(text, VALUE_PATTERNS):
            label = "值回票价"
        elif feature == "场景":
            label = "灵感打开"
        else:
            label = "满意认可"
    elif polarity == "负向":
        if contains_any(text, NOISE_PATTERNS):
            label = "被噪音打断"
        elif contains_any(text, CONTROL_PATTERNS):
            label = "失控难调"
        elif contains_any(text, CONNECT_PATTERNS):
            label = "焦虑不确定"
        elif viewpoint in {"期望落差", "对比"} or contains_any(text, EXPECTATION_PATTERNS):
            label = "失望落差"
        elif intensity == "强烈":
            label = "愤怒拒绝"
        else:
            label = "烦躁麻烦"
    else:
        if viewpoint == "建议":
            label = "理性期待"
        elif viewpoint == "对比":
            label = "理性对比"
        else:
            label = "观察确认"

    trigger = feature
    if label in {"被噪音打断"}:
        trigger = "音质"
    elif label in {"失控难调"}:
        trigger = "音量"
    elif label in {"焦虑不确定"}:
        trigger = "连接"

    return polarity, intensity, label, trigger


def classify_scene(text: str) -> str:
    return first_match(text, SCENE_RULES, "通用补洞")


def build_scene_card(scene_tag: str, feature: str, need: str):
    persona, place, motivation, chain = SCENE_TEMPLATES.get(scene_tag, SCENE_TEMPLATES["通用补洞"])
    chain = f"{chain}，核心希望是通过{feature}获得{need}。"
    return persona, place, motivation, chain


def classify_decision_signal(viewpoint: str, polarity: str) -> str:
    if viewpoint == "疑似灌水":
        return "噪声样本"
    if viewpoint in {"建议"}:
        return "合理增配"
    if viewpoint in {"抱怨", "缺陷复现", "期望落差"} or polarity == "负向":
        return "放弃采用"
    if viewpoint in {"优点", "对比"} and polarity == "正向":
        return "喜欢什么"
    return "待归纳"


def aha_scores(text: str, viewpoint: str, decision_signal: str, polarity: str, scene_tag: str, need: str, image_count: int):
    if polarity == "负向" and viewpoint not in {"建议", "对比"}:
        return "否", 0, "低", "低"
    if decision_signal not in {"喜欢什么", "合理增配"} and viewpoint not in {"对比"}:
        return "否", 0, "低", "低"
    score = 0
    if contains_any(text, STRONG_POSITIVE):
        score += 2
    if any(phrase in text for phrase in ("i use this to", "finally", "exactly what i needed", "正是我需要的", "终于", "我用它来")):
        score += 2
    if scene_tag != "通用补洞":
        score += 1
    if need in {"降摩擦", "系统补足", "灵感启发", "场景适配"}:
        score += 1
    if viewpoint in {"对比", "建议"}:
        score += 1
    if image_count > 0:
        score += 1
    if len(normalize_key(text)) >= 50:
        score += 1
    label = "是" if score >= 4 else "否"
    relation = "高" if score >= 6 else "中" if score >= 4 else "低"
    originality = "高" if score >= 7 or need in {"灵感启发", "系统补足"} else "中" if score >= 4 else "低"
    return label, score, relation, originality


def build_opportunity(decision_signal: str, feature: str, need: str, scene_tag: str):
    if decision_signal == "喜欢什么":
        pain = f"用户原本在 {scene_tag} 中要用额外步骤解决 {need}，该产品把这段摩擦显著降低。"
    elif decision_signal == "放弃采用":
        pain = f"用户在 {feature} 上无法稳定获得 {need}，因此宁可放弃或更换方案。"
    elif decision_signal == "合理增配":
        pain = f"用户已认可主链路，但希望在 {feature} 上进一步补足 {need}。"
    else:
        pain = f"该评论仍与 {feature} 和 {need} 相关，需要结合上下文再判断决策意义。"
    hint = f"产品定义可优先强化 {feature}，面向 {scene_tag} 用户把 {need} 做成更明确的能力承诺。"
    analysis = f"{pain} {hint}"
    return pain, hint, analysis


def ranking_score(viewpoint: str, polarity: str, intensity: str, image_count: int, text: str) -> int:
    score = 0
    score += {"优点": 2, "抱怨": 3, "建议": 4, "对比": 3, "期望落差": 4, "缺陷复现": 5, "疑似灌水": 1, "观点其他": 1}.get(viewpoint, 1)
    score += {"正向": 1, "中性": 0, "负向": 2}.get(polarity, 0)
    score += {"轻微": 0, "中等": 1, "强烈": 2}.get(intensity, 0)
    score += 1 if image_count > 0 else 0
    score += 2 if len(normalize_key(text)) >= 80 else 1 if len(normalize_key(text)) >= 40 else 0
    return score


def copy_sheet(source_ws, output_wb, title: str):
    copied = output_wb.create_sheet(title)
    for row in source_ws.iter_rows():
        copied.append([cell.value for cell in row])
    copied.freeze_panes = "A2"
    return copied


def main():
    parser = argparse.ArgumentParser(description="Apply first-pass VOC tags to cleaned review rows.")
    parser.add_argument("workbook", help="Path to workbook containing CleanedComments")
    parser.add_argument("--sheet", default="CleanedComments", help="Input sheet name")
    parser.add_argument("--output", help="Output xlsx path")
    args = parser.parse_args()

    workbook_path = Path(args.workbook).resolve()
    wb = load_workbook(workbook_path, data_only=True)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet '{args.sheet}' not found in {workbook_path}")
    source_ws = wb[args.sheet]

    rows = []
    for row in row_dicts(source_ws):
        if str(row.get("is_valid_feedback", "")).upper() != "TRUE":
            continue
        comment_text = normalize_text(str(row.get("cleaned_comment", "")))
        source_theme = source_theme_text(row)
        context_text = normalize_text(
            " ".join(
                [
                    str(row.get("source_primary_theme", "")),
                    str(row.get("source_secondary_theme", "")),
                    str(row.get("source_topic_labels", "")),
                    str(row.get("product_name", "")),
                    str(row.get("scene_name", "")),
                    str(row.get("keyword_source", "")),
                    str(row.get("cleaned_comment", "")),
                ]
            )
        )
        rating = parse_rating(row.get("rating"))
        viewpoint = classify_viewpoint(comment_text, rating)
        feature = classify_feature(comment_text)
        if feature == "基础":
            source_feature = classify_feature_from_source(source_theme)
            feature = source_feature or classify_feature(context_text)
        secondary_need = classify_secondary_need(comment_text or context_text, feature, source_theme)
        need = classify_hidden_need(comment_text, feature)
        polarity, intensity, emotion_label, emotion_trigger = classify_emotion(comment_text, rating, viewpoint, feature)
        scene_tag = classify_scene(context_text)
        scene_persona, scene_place, scene_motivation, scene_chain = build_scene_card(scene_tag, feature, need)
        decision_signal = classify_decision_signal(viewpoint, polarity)
        image_count = int(parse_rating(row.get("image_count")))
        aha_label, aha_score, aha_relation, aha_originality = aha_scores(
            comment_text, viewpoint, decision_signal, polarity, scene_tag, need, image_count
        )
        pain_point, product_hint, aha_analysis = build_opportunity(decision_signal, feature, need, scene_tag)
        rank_score = ranking_score(viewpoint, polarity, intensity, image_count, comment_text)

        tagged = {column: row.get(column, "") for column in TAG_COLUMNS if column in row}
        tagged.update(
            {
                "观点类型": viewpoint,
                "一级功能": feature,
                "二级功能需求": secondary_need,
                "底层需求": need,
                "决策信号": decision_signal,
                "情绪极性": polarity,
                "情绪强度": intensity,
                "情绪标签": emotion_label,
                "情绪触发点": emotion_trigger,
                "场景标签": scene_tag,
                "场景人物": scene_persona,
                "场景地点": scene_place,
                "场景动机": scene_motivation,
                "场景链路": scene_chain,
                "嘿哈时刻": aha_label,
                "嘿哈分数": aha_score,
                "嘿哈关联度": aha_relation,
                "嘿哈独到性": aha_originality,
                "痛点机会": pain_point,
                "产品定义启发": product_hint,
                "嘿哈分析": aha_analysis,
                "排序分": rank_score,
            }
        )
        rows.append(tagged)

    output_path = (
        Path(args.output).resolve()
        if args.output
        else workbook_path.with_name("voc_tagged_comments.xlsx")
    )

    out_wb = Workbook()
    out_wb.remove(out_wb.active)
    copy_sheet(source_ws, out_wb, "CleanedComments")
    tagged_ws = out_wb.create_sheet("TaggedComments")
    tagged_ws.append(TAG_COLUMNS)
    for row in rows:
        tagged_ws.append([row.get(column, "") for column in TAG_COLUMNS])
    tagged_ws.freeze_panes = "A2"

    if "DroppedRows" in wb.sheetnames:
        copy_sheet(wb["DroppedRows"], out_wb, "DroppedRows")
    if "Metadata" in wb.sheetnames:
        copy_sheet(wb["Metadata"], out_wb, "Metadata")

    out_wb.save(output_path)
    print(json.dumps({"output": str(output_path), "tagged_rows": len(rows)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
