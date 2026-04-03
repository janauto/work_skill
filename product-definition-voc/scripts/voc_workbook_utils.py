#!/usr/bin/env python3
from __future__ import annotations

import re
from collections import Counter
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook


HEADER_ALIASES = {
    "asin": ["asin", "sku id", "商品id", "listing asin"],
    "product_name": [
        "产品",
        "产品名",
        "商品标题",
        "商品名称",
        "product_name",
        "listing title",
        "product title",
        "product",
        "sku",
        "型号",
        "机型",
    ],
    "product_image": ["商品图片", "product image", "image url", "主图", "主图链接", "cover image"],
    "product_link": ["商品链接", "product link", "listing url", "url", "link"],
    "thread_url": ["thread_url", "thread url", "讨论链接", "帖子链接"],
    "source_primary_theme": ["primary_theme", "primary theme", "一级功", "一级功能", "主话题"],
    "source_secondary_theme": ["secondary_themes", "secondary_theme", "secondary themes", "二级功", "二级功能", "二级话题"],
    "source_topic_labels": ["topic_labels", "topic label", "topic labels", "话题标签", "标签"],
    "scene_name": ["场景名称", "场景", "scene", "use case", "usage scene"],
    "keyword_source": ["关键词来源", "关键词", "keyword", "search keyword", "query"],
    "store": ["店铺", "store", "shop", "seller", "merchant"],
    "country": ["国家", "country", "market", "站点"],
    "rating": ["评分", "rating", "stars", "星级", "score"],
    "rating_text": ["评分文本", "rating text"],
    "review_date": ["评论时间", "日期", "date", "time", "review date", "post_date"],
    "raw_title": ["评论标题", "review title", "headline", "标题", "thread_title"],
    "translated_title": ["评论标题中文", "标题中文", "translated title", "title cn"],
    "raw_comment": [
        "评论内容",
        "原文",
        "英文评论",
        "review",
        "comment",
        "review text",
        "comment text",
        "body",
        "post_text",
        "raw_text",
        "买家备注",
    ],
    "translated_comment": ["评论内容中文", "中文翻译", "翻译", "译文", "comment cn", "review cn"],
    "image_count": ["图片数量", "image count", "photo count", "image num", "image_count"],
}

IMAGE_HEADER_PATTERN = re.compile(r"^(图片\d+|image\d+|image \d+|photo\d+|photo \d+|image_urls?)$", re.IGNORECASE)


def load_workbook_safe(path: str | Path, data_only: bool = True):
    return load_workbook(Path(path), data_only=data_only)


def normalize_header(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value).strip().lower())


def normalize_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_key(value) -> str:
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", normalize_text(value))


def display_text(value) -> str:
    return str(value).strip() if value is not None else ""


def detect_header_row(ws, scan_rows: int = 8) -> Tuple[int, Dict[str, int], List[int]]:
    best_row = 1
    best_score = -1
    best_map: Dict[str, int] = {}
    best_images: List[int] = []

    for row_idx in range(1, min(scan_rows, ws.max_row) + 1):
        headers = {
            col_idx: normalize_header(ws.cell(row_idx, col_idx).value)
            for col_idx in range(1, ws.max_column + 1)
        }
        column_map: Dict[str, int] = {}
        image_columns: List[int] = []
        score = 0

        for semantic, aliases in HEADER_ALIASES.items():
            alias_set = {normalize_header(alias) for alias in aliases}
            for col_idx, header in headers.items():
                if header and header in alias_set:
                    column_map[semantic] = col_idx
                    score += 2 if semantic in {"raw_comment", "translated_comment", "product_name"} else 1
                    break

        for col_idx, header in headers.items():
            if not header:
                continue
            if IMAGE_HEADER_PATTERN.match(header) or header in {"图片路径", "图片目录", "imagepath", "image url", "photo"}:
                image_columns.append(col_idx)
                score += 1

        if score > best_score:
            best_row = row_idx
            best_score = score
            best_map = column_map
            best_images = image_columns

    return best_row, best_map, best_images


def infer_source_type(sheet_name: str, column_map: Dict[str, int]) -> str:
    lowered = sheet_name.lower()
    if "review" in lowered or "评论" in sheet_name:
        return "review"
    if "feedback" in lowered or "反馈" in sheet_name:
        return "feedback"
    if any(key in column_map for key in ("raw_comment", "translated_comment")):
        return "review"
    return "unknown"


def is_candidate_sheet(column_map: Dict[str, int]) -> bool:
    return any(name in column_map for name in ("raw_comment", "translated_comment", "raw_title", "translated_title"))


def has_text_columns(column_map: Dict[str, int]) -> bool:
    return any(name in column_map for name in ("raw_comment", "translated_comment", "raw_title", "translated_title"))


def sheet_profile(ws, workbook_path: str | Path) -> Dict[str, object]:
    header_row, column_map, image_columns = detect_header_row(ws)
    return {
        "sheet": ws.title,
        "header_row": header_row,
        "column_map": column_map,
        "image_columns": image_columns,
        "source_type": infer_source_type(ws.title, column_map),
        "candidate": is_candidate_sheet(column_map),
        "workbook": str(Path(workbook_path).resolve()),
    }


def row_value(ws, row_idx: int, column_map: Dict[str, int], semantic: str) -> str:
    column_idx = column_map.get(semantic)
    if not column_idx:
        return ""
    return display_text(ws.cell(row_idx, column_idx).value)


def choose_comment_text(record: Dict[str, str]) -> str:
    if record.get("translated_title") or record.get("translated_comment"):
        parts = [record.get("translated_title", ""), record.get("translated_comment", "")]
        return "。".join(part for part in parts if part).strip("。")
    if record.get("raw_title") or record.get("raw_comment"):
        parts = [record.get("raw_title", ""), record.get("raw_comment", "")]
        return "。".join(part for part in parts if part).strip("。")
    return ""


def build_image_refs(ws, row_idx: int, image_columns: List[int]) -> Tuple[int, str]:
    refs: List[str] = []
    for column_idx in image_columns:
        value = display_text(ws.cell(row_idx, column_idx).value)
        if value:
            refs.append(value)
    return len(refs), " | ".join(refs)


def focus_matches_record(focus: Optional[str], record: Dict[str, str]) -> bool:
    if not focus:
        return True
    target = normalize_key(focus)
    if not target:
        return True
    combined = " ".join(
        [
            record.get("product_name", ""),
            record.get("scene_name", ""),
            record.get("keyword_source", ""),
            record.get("source_sheet", ""),
            record.get("raw_title", ""),
            record.get("translated_title", ""),
            record.get("raw_comment", ""),
            record.get("translated_comment", ""),
        ]
    )
    return target in normalize_key(combined) or normalize_text(focus) in normalize_text(combined)


def collect_focus_candidates(path: str | Path) -> List[str]:
    workbook_path = Path(path)
    wb = load_workbook_safe(workbook_path, data_only=True)
    candidates: Counter[str] = Counter()
    for ws in wb.worksheets:
        profile = sheet_profile(ws, workbook_path)
        if not profile["candidate"]:
            continue
        header_row = int(profile["header_row"])
        column_map = profile["column_map"]
        for semantic in ("product_name", "scene_name", "keyword_source"):
            column_idx = column_map.get(semantic)
            if not column_idx:
                continue
            for row_idx in range(header_row + 1, min(ws.max_row, header_row + 120) + 1):
                value = display_text(ws.cell(row_idx, column_idx).value)
                if value:
                    candidates[value] += 1
        candidates[ws.title] += 1

    return [name for name, _ in candidates.most_common(20)]


def build_product_lookup(workbook_path: str | Path):
    wb = load_workbook_safe(workbook_path, data_only=True)
    lookup: dict[tuple[str, str], dict[str, str]] = {}
    for ws in wb.worksheets:
        header_row, column_map, _image_columns = detect_header_row(ws)
        if "product_image" not in column_map and "product_link" not in column_map:
            continue
        for row_idx in range(header_row + 1, ws.max_row + 1):
            product_name = row_value(ws, row_idx, column_map, "product_name")
            scene_name = row_value(ws, row_idx, column_map, "scene_name")
            keyword_source = row_value(ws, row_idx, column_map, "keyword_source")
            asin = row_value(ws, row_idx, column_map, "asin")
            image_value = row_value(ws, row_idx, column_map, "product_image")
            link_value = row_value(ws, row_idx, column_map, "product_link")
            if not any((image_value, link_value)):
                continue
            payload = {"product_image": image_value, "product_link": link_value}
            for kind, value in (
                ("asin", asin),
                ("product_name", product_name),
                ("scene_name", scene_name),
                ("keyword_source", keyword_source),
            ):
                key = (kind, normalize_key(value))
                if value and key not in lookup:
                    lookup[key] = payload
    return lookup


def collect_rows(path: str | Path, focus: Optional[str] = None, sheet_name: Optional[str] = None):
    workbook_path = Path(path).resolve()
    wb = load_workbook_safe(workbook_path, data_only=True)
    product_lookup = build_product_lookup(workbook_path)
    collected = []
    profiles = []

    for ws in wb.worksheets:
        profile = sheet_profile(ws, workbook_path)
        profiles.append(profile)
        if sheet_name and ws.title != sheet_name:
            continue
        if not profile["candidate"]:
            continue
        if not has_text_columns(profile["column_map"]):
            continue

        header_row = int(profile["header_row"])
        column_map = profile["column_map"]
        image_columns = profile["image_columns"]

        for row_idx in range(header_row + 1, ws.max_row + 1):
            record = {
                "record_id": f"{ws.title}-{row_idx}",
                "focus_name": focus or "",
                "asin": row_value(ws, row_idx, column_map, "asin"),
                "product_name": row_value(ws, row_idx, column_map, "product_name"),
                "product_image": row_value(ws, row_idx, column_map, "product_image"),
                "product_link": row_value(ws, row_idx, column_map, "product_link"),
                "thread_url": row_value(ws, row_idx, column_map, "thread_url"),
                "source_primary_theme": row_value(ws, row_idx, column_map, "source_primary_theme"),
                "source_secondary_theme": row_value(ws, row_idx, column_map, "source_secondary_theme"),
                "source_topic_labels": row_value(ws, row_idx, column_map, "source_topic_labels"),
                "scene_name": row_value(ws, row_idx, column_map, "scene_name"),
                "keyword_source": row_value(ws, row_idx, column_map, "keyword_source"),
                "source_type": profile["source_type"],
                "source_sheet": ws.title,
                "source_row": row_idx,
                "store": row_value(ws, row_idx, column_map, "store") or "未提供",
                "country": row_value(ws, row_idx, column_map, "country") or "未提供",
                "rating": row_value(ws, row_idx, column_map, "rating"),
                "rating_text": row_value(ws, row_idx, column_map, "rating_text"),
                "review_date": row_value(ws, row_idx, column_map, "review_date"),
                "raw_title": row_value(ws, row_idx, column_map, "raw_title"),
                "translated_title": row_value(ws, row_idx, column_map, "translated_title"),
                "raw_comment": row_value(ws, row_idx, column_map, "raw_comment"),
                "translated_comment": row_value(ws, row_idx, column_map, "translated_comment"),
            }
            record["cleaned_comment"] = choose_comment_text(record)
            image_count, image_refs = build_image_refs(ws, row_idx, image_columns)
            record["image_count"] = str(image_count)
            record["image_refs"] = image_refs
            if not record["product_image"]:
                for kind, value in (
                    ("asin", record["asin"]),
                    ("product_name", record["product_name"]),
                    ("scene_name", record["scene_name"]),
                    ("keyword_source", record["keyword_source"]),
                ):
                    key = (kind, normalize_key(value))
                    if value and key in product_lookup:
                        record["product_image"] = product_lookup[key].get("product_image", "")
                        record["product_link"] = record["product_link"] or product_lookup[key].get("product_link", "")
                        break
            if focus_matches_record(focus, record):
                collected.append(record)

    return collected, profiles
